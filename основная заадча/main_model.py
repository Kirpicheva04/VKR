import tkinter as tk
from tkinter import ttk, messagebox
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from mpl_toolkits.mplot3d import Axes3D
import threading
import datetime
import os
import tkinter as tk
from tkinter import ttk, messagebox
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from scipy.optimize import fsolve
from scipy.integrate import odeint
import warnings
import datetime
import os
from io import StringIO


warnings.filterwarnings('ignore')


class ActivatorInhibitorApp:
    def __init__(self, root):
        self.precision = 16  # точность вычислений
        self.root = root
        self.root.title("Анализ модели Майнхардта")
        self.root.geometry("1400x900")
        self.root.configure(bg='#e6f2ff')
        self.refined_equilibrium_points = []  # Хранилище уточненных точек (a,b,c)
        self.equilibrium_points_found = False  # Флаг, что точки найдены
        self.graphic_equilibrium_points = []  # Хранилище графических точек (a,c) для метода Ньютона

        self.params = {
            'pa': 1.0, 'ka': 5.0, 'ga': 0.01,  # ga, а не sigma_a
            'pb': 1.0, 'kb': 5.0, 'gb': 0.01,  # gb, а не sigma_b
            'pc': 0.1, 'Da': 0.5, 'Db': 0.5, 'Dc': 0.5
        }

        # Стили
        self.setup_styles()
        self.setup_gui()

    def setup_styles(self):
        """Настройка стилей интерфейса"""
        self.style = ttk.Style()
        self.style.configure('Blue.TFrame', background='#e6f2ff')
        self.style.configure('Blue.TLabelframe', background='#cce5ff', foreground='#0066cc')
        self.style.configure('Blue.TLabelframe.Label', background='#cce5ff', foreground='#0066cc')
        self.style.configure('Title.TLabel', background='#e6f2ff', foreground='#0066cc',
                             font=('Arial', 14, 'bold'))
        self.style.configure('Param.TLabel', background='#cce5ff', foreground='#0066cc')

    def setup_gui(self):
        """Создание интерфейса с вкладками"""
        # Главный заголовок
        title_label = ttk.Label(self.root, text="АНАЛИЗ МОДЕЛИ МАЙНХАРДТА",
                                style='Title.TLabel')
        title_label.pack(pady=10)

        # Создаем вкладки
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Вкладка 1: Аналитическое исследование (старая)
        self.tab1 = ttk.Frame(self.notebook, style='Blue.TFrame')
        self.notebook.add(self.tab1, text="Аналитическое исследование")

        # Вкладка 2: Численное решение (новая)
        self.tab2 = ttk.Frame(self.notebook, style='Blue.TFrame')
        self.notebook.add(self.tab2, text="Численное решение")

        # Настраиваем содержимое первой вкладки
        self.setup_original_gui_tab1()

        # Создаем вторую вкладку
        self.numerical_tab = NumericalTab(self.tab2, self)

    def setup_original_gui_tab1(self):
        """Переносим сюда весь код из старого setup_gui для первой вкладки"""
        # Основной контейнер
        main_container = ttk.Frame(self.tab1, style='Blue.TFrame')
        main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Левая панель - управление
        left_panel = ttk.Frame(main_container, style='Blue.TFrame', width=350)
        left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        left_panel.pack_propagate(False)

        # Правая панель - графика и информация о задаче
        right_panel = ttk.Frame(main_container, style='Blue.TFrame')
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # Создаем компоненты (копируем из старого setup_gui)
        self.setup_control_panel(left_panel)
        self.setup_right_panel(right_panel)

    def setup_control_panel(self, parent):
        """Левая панель - параметры, визуализация, история"""
        # Параметры системы
        params_frame = ttk.LabelFrame(parent, text="ПАРАМЕТРЫ МОДЕЛИ", style='Blue.TLabelframe')
        params_frame.pack(fill=tk.X, pady=(0, 10))

        # Создаем контейнер с двумя колонками
        params_container = ttk.Frame(params_frame, style='Blue.TFrame')
        params_container.pack(fill=tk.X, padx=5, pady=5)

        # ЛЕВАЯ КОЛОНКА
        left_column = ttk.Frame(params_container, style='Blue.TFrame')
        left_column.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

        # ПРАВАЯ КОЛОНКА
        right_column = ttk.Frame(params_container, style='Blue.TFrame')
        right_column.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.param_entries = {}

        # === ЛЕВАЯ КОЛОНКА: 5 параметров ===
        left_params = ['pa', 'pb', 'pc', 'ka', 'kb']

        for param in left_params:
            param_frame = ttk.Frame(left_column, style='Blue.TFrame')
            param_frame.pack(fill=tk.X, pady=3)

            # Отображаемое имя с индексами
            display_name = {
                'pa': 'ρ_a',
                'pb': 'ρ_b',
                'pc': 'ρ_c',
                'ka': 'k_a',
                'kb': 'k_b'
            }.get(param, param)

            label = ttk.Label(param_frame, text=f"{display_name} =", width=8, style='Param.TLabel')
            label.pack(side=tk.LEFT)

            # Значения по умолчанию
            default_val = 1.0 if param in ['pa', 'pb', 'pc', 'ka', 'kb'] else 0.1
            if param not in self.params:
                self.params[param] = default_val

            entry = ttk.Entry(param_frame, width=12)
            entry.insert(0, str(self.params[param]))
            entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

            self.param_entries[param] = entry

        # === ПРАВАЯ КОЛОНКА: 5 параметров ===
        right_params = ['ga', 'gb', 'Da', 'Db', 'Dc']  # ga и gb, а не sigma_a, sigma_b

        for param in right_params:
            param_frame = ttk.Frame(right_column, style='Blue.TFrame')
            param_frame.pack(fill=tk.X, pady=3)

            # Отображаемое имя с индексами
            display_name = {
                'ga': 'σ_a',
                'gb': 'σ_b',
                'Da': 'D_a',
                'Db': 'D_b',
                'Dc': 'D_c'
            }.get(param, param)

            label = ttk.Label(param_frame, text=f"{display_name} =", width=8, style='Param.TLabel')
            label.pack(side=tk.LEFT)

            # Значения по умолчанию
            default_val = 0.1
            if param not in self.params:
                self.params[param] = default_val

            entry = ttk.Entry(param_frame, width=12)
            entry.insert(0, str(self.params[param]))
            entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

            self.param_entries[param] = entry

        # Кнопка применения параметров
        apply_button = tk.Button(params_frame, text="ПРИМЕНИТЬ ПАРАМЕТРЫ",
                                 command=self.apply_parameters,
                                 bg='#4da6ff', fg='white',
                                 font=('Arial', 10, 'bold'),
                                 relief=tk.RAISED, bd=2)
        apply_button.pack(fill=tk.X, padx=5, pady=10)

        # ВИЗУАЛИЗАЦИЯ
        viz_frame = ttk.LabelFrame(parent, text="ВИЗУАЛИЗАЦИЯ", style='Blue.TLabelframe')
        viz_frame.pack(fill=tk.X, pady=(0, 10))

        # Кнопки визуализации
        viz_buttons_frame = ttk.Frame(viz_frame, style='Blue.TFrame')
        viz_buttons_frame.pack(fill=tk.X, padx=5, pady=5)

        # Кнопка "Точки равновесия" (заменяет фазовый портрет)
        self.equilibrium_button = tk.Button(viz_buttons_frame, text="ТОЧКИ РАВНОВЕСИЯ",
                                            command=self.plot_equilibrium_points,
                                            bg='#66b3ff', fg='white',
                                            font=('Arial', 9, 'bold'),
                                            width=30, height=1)
        self.equilibrium_button.pack(pady=2)

        # Кнопка "Метод Ньютона" (заменяет бифуркационную диаграмму)
        self.neumann_button = tk.Button(viz_buttons_frame, text="МЕТОД НЬЮТОНА",
                                        command=self.plot_neumann_table,
                                        bg='#66b3ff', fg='white',
                                        font=('Arial', 9, 'bold'),
                                        width=30, height=1)
        self.neumann_button.pack(pady=2)

        # АНАЛИЗ ТЬЮРИНГА
        self.turing_button = tk.Button(viz_buttons_frame, text="НЕУСТОЙЧИВОСТЬ ТЬЮРИНГА",
                                       command=self.plot_turing_analysis,
                                       bg='#66b3ff', fg='white',
                                       font=('Arial', 9, 'bold'),
                                       width=30, height=1)
        self.turing_button.pack(pady=2)


        # Кнопка вывода данных в файл
        export_frame = ttk.Frame(parent, style='Blue.TFrame')
        export_frame.pack(fill=tk.X, pady=(0, 5))

        self.export_button = tk.Button(export_frame, text="ВЫВОД ДАННЫХ В ФАЙЛ",
                                       command=self.export_to_file,
                                       bg='#4da6ff', fg='white',
                                       font=('Arial', 10, 'bold'),
                                       relief=tk.RAISED, bd=3)
        self.export_button.pack(fill=tk.X, pady=5)

        # ИСТОРИЯ РАСЧЕТОВ
        calc_frame = ttk.LabelFrame(parent, text="ИСТОРИЯ РАСЧЕТОВ", style='Blue.TLabelframe')
        calc_frame.pack(fill=tk.BOTH, expand=True)

        self.calc_text = tk.Text(calc_frame, height=25, width=40,
                                 bg='#f0f8ff', fg='#0066cc',
                                 font=('Arial', 9))
        scrollbar = ttk.Scrollbar(calc_frame, command=self.calc_text.yview)
        self.calc_text.configure(yscrollcommand=scrollbar.set)

        self.calc_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Статус бар
        self.status_var = tk.StringVar(value="Готов к работе")
        status_label = tk.Label(parent, textvariable=self.status_var,
                                bg='#cce5ff', fg='#0066cc', relief=tk.SUNKEN,
                                font=('Arial', 9))
        status_label.pack(fill=tk.X, pady=(10, 0))

    def show_parameters_applied_message(self):
        """Показывает информационные рамки при применении параметров"""
        self.fig.clear()

        # Верхняя рамка
        ax1 = self.fig.add_subplot(211)
        ax1.set_facecolor('#f0f8ff')

        # Делаем окантовку графика цветной
        for spine in ax1.spines.values():
            spine.set_edgecolor('#0066cc')
            spine.set_linewidth(1.5)

        ax1.text(0.5, 0.6, 'ПАРАМЕТРЫ МОДЕЛИ ПРИМЕНЕНЫ',
                 horizontalalignment='center', verticalalignment='center',
                 transform=ax1.transAxes, fontsize=14, fontweight='bold',
                 color='#0066cc')

        params_text = (f"ρ_a = {self.params['pa']:.3f}    k_a = {self.params['ka']:.3f}    σ_a = {self.params['ga']:.3f}\n"
                       f"ρ_b = {self.params['pb']:.3f}    k_b = {self.params['kb']:.3f}    σ_b = {self.params['gb']:.3f}\n"
                       f"ρ_c = {self.params['pc']:.3f}    D_a = {self.params['Da']:.3f}    D_b = {self.params['Db']:.3f}    Dc = {self.params['Dc']:.3f}")

        ax1.text(0.5, 0.3, params_text,
                 horizontalalignment='center', verticalalignment='center',
                 transform=ax1.transAxes, fontsize=11,
                 color='#0066cc')

        ax1.set_xticks([])
        ax1.set_yticks([])

        # Нижняя рамка
        ax2 = self.fig.add_subplot(212)
        ax2.set_facecolor('#f0f8ff')

        # Делаем окантовку графика цветной
        for spine in ax2.spines.values():
            spine.set_edgecolor('#0066cc')
            spine.set_linewidth(1.5)

        ax2.text(0.5, 0.5, 'Для построения графиков решения выберите команду из раздела "ВИЗУАЛИЗАЦИЯ"',
                 horizontalalignment='center', verticalalignment='center',
                 transform=ax2.transAxes, fontsize=12,
                 color='#0066cc')

        ax2.set_xticks([])
        ax2.set_yticks([])

        self.fig.subplots_adjust(hspace=0.3)
        self.canvas.draw()

    def setup_right_panel(self, parent):
        """Правая панель - постановка задачи и графики"""
        # Верхняя часть - постановка задачи
        top_panel = ttk.Frame(parent, style='Blue.TFrame')
        top_panel.pack(fill=tk.X, pady=(0, 10))

        # Постановка задачи
        info_frame = ttk.LabelFrame(top_panel, text="ПОСТАНОВКА ЗАДАЧИ", style='Blue.TLabelframe')
        info_frame.pack(fill=tk.BOTH, expand=True, padx=(0, 0))

        # Текст задачи
        problem_text = (
            "СИСТЕМА УРАВНЕНИЙ:\n"
            "∂a/∂t = ρ_a (c/(1 + k_a·b²) - a) + σ_a + D_a·Δ²a\n"
            "∂b/∂t = ρ_b (1/(1 + k_b·a²·c) - b) + σ_b + D_b·Δ²b\n"
            "∂c/∂t = ρ_c (b - c·a) + D_c·Δ²c\n\n"
            "ГДЕ:\n"
            "  a, b, c – концентрации веществ\n"
            "  ρ_a, ρ_b, ρ_c – коэффициенты скорости реакций\n"
            "  k_a, k_b – коэффициенты насыщения\n"
            "  σ_a, σ_b – базовые скорости реакции (подкачка)\n"
            "  D_a, D_b, D_c – коэффициенты диффузии"
        )

        # Текстовое поле с прокруткой
        self.info_text = tk.Text(info_frame, height=10, width=80,
                                 bg='#f0f8ff', fg='#0066cc',
                                 font=('Courier', 10), wrap=tk.WORD)
        scrollbar_info = ttk.Scrollbar(info_frame, command=self.info_text.yview)
        self.info_text.configure(yscrollcommand=scrollbar_info.set)

        self.info_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar_info.pack(side=tk.RIGHT, fill=tk.Y)

        # Вставляем текст
        self.info_text.insert(1.0, problem_text)
        self.info_text.config(state=tk.DISABLED)

        # Нижняя часть - графики
        bottom_panel = ttk.Frame(parent, style='Blue.TFrame')
        bottom_panel.pack(fill=tk.BOTH, expand=True)

        # Область для графиков
        plot_frame = ttk.LabelFrame(bottom_panel, text="ГРАФИКИ РЕШЕНИЯ", style='Blue.TLabelframe')
        plot_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Контейнер для графика
        graph_container = ttk.Frame(plot_frame, style='Blue.TFrame')
        graph_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Создаем фигуру matplotlib
        self.fig = Figure(figsize=(10, 7), facecolor='white')
        self.canvas = FigureCanvasTkAgg(self.fig, graph_container)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        # Начальный график с информацией
        self.show_welcome_plot()

    def show_welcome_plot(self):
        """Показывает приветственный график"""
        self.fig.clear()
        ax = self.fig.add_subplot(111)
        ax.set_facecolor('#f0f8ff')

        info_text = 'Добро пожаловать!\n\n' \
                    'Для начала работы:\n' \
                    '1. Задайте параметры модели\n' \
                    '2. Нажмите "Применить параметры"\n' \
                    '3. Используйте кнопки визуализации\n'

        ax.text(0.5, 0.5, info_text, horizontalalignment='center',
                verticalalignment='center', transform=ax.transAxes,
                fontsize=12, color='#0066cc')
        ax.set_xticks([])
        ax.set_yticks([])
        ax.set_title('Анализ модели Майнхардта',
                     color='#0066cc', fontsize=14, pad=20)

        self.canvas.draw()

    # ====================== МЕТОДЫ СИСТЕМЫ ======================

    def system_eq(self, vars):
        """Система уравнений """
        a, c = vars
        pa, ka, ga, pb, kb, gb = list(self.params.values())
        eq1 = pa * (c / (1 + ka * a ** 2 * c ** 2) - a) + ga
        eq2 = pb * (1 / (1 + kb * a ** 2 * c) - a * c) + gb
        return [eq1, eq2]

    def system_ode(self, vars, t):
        """Для интегрирования ODE"""
        a, c = vars

        # Явно берем только нужные параметры
        pa = self.params['pa']
        ka = self.params['ka']
        ga = self.params['ga']
        pb = self.params['pb']
        kb = self.params['kb']
        gb = self.params['gb']

        dadt = pa * (c / (1 + ka * a ** 2 * c ** 2) - a) + ga
        dcdt = pb * (1 / (1 + kb * a ** 2 * c) - a * c) + gb
        return [dadt, dcdt]

    def jacobian(self, point):
        """Матрица Якоби системы"""
        a, c = point

        # Явно берем только нужные параметры
        pa = self.params['pa']
        ka = self.params['ka']
        ga = self.params['ga']
        pb = self.params['pb']
        kb = self.params['kb']
        gb = self.params['gb']

        denominator1 = (1 + ka * a ** 2 * c ** 2) ** 2
        df1_da = pa * ((-2 * ka * a * c ** 3) / denominator1 - 1)
        numerator1 = (1 + ka * a ** 2 * c ** 2) - 2 * ka * a ** 2 * c ** 2
        df1_dc = pa * (numerator1 / denominator1)

        denominator2 = (1 + kb * a ** 2 * c) ** 2
        df2_da = pb * ((-2 * kb * a * c) / denominator2 - c)
        df2_dc = pb * ((-kb * a ** 2) / denominator2 - a)

        return np.array([[df1_da, df1_dc],
                         [df2_da, df2_dc]])

    def find_equilibrium_points(self):
        """Находит точки равновесия системы"""
        solutions = []

        # Явно берем только нужные параметры
        pa = self.params['pa']
        ka = self.params['ka']
        ga = self.params['ga']
        pb = self.params['pb']
        kb = self.params['kb']
        gb = self.params['gb']

        # Пробуем разные начальные приближения
        initial_guesses = []
        for a0 in np.linspace(0.1, 5.0, 30):
            for c0 in np.linspace(0.1, 5.0, 30):
                initial_guesses.append([a0, c0])

        for guess in initial_guesses:
            try:
                # Используем lambda с явными параметрами
                def eq_func(vars):
                    a, c = vars
                    eq1 = pa * (c / (1 + ka * a ** 2 * c ** 2) - a) + ga
                    eq2 = pb * (1 / (1 + kb * a ** 2 * c) - a * c) + gb
                    return [eq1, eq2]

                sol = fsolve(eq_func, guess, maxfev=2000, xtol=1e-10)
                a, c = sol

                # Проверяем, что это действительно решение
                f1, f2 = eq_func(sol)
                if abs(f1) < 1e-8 and abs(f2) < 1e-8 and a > 0 and c > 0:
                    # Проверяем уникальность
                    unique = True
                    for existing in solutions:
                        if np.linalg.norm(existing - sol) < 1e-6:
                            unique = False
                            break
                    if unique:
                        solutions.append(sol)
            except:
                continue

        return solutions

    def analyze_stability(self, point):
        """Анализ устойчивости точки"""
        J = self.jacobian(point)
        tr = np.trace(J)
        det = np.linalg.det(J)
        eigvals = np.linalg.eigvals(J)

        # Дискриминант: D = tr² - 4det
        discriminant = tr ** 2 - 4 * det

        # Определяем тип точки
        if det > 0 and tr < 0:
            if discriminant < 0:
                stability = "устойчивый фокус"
            elif discriminant > 0:
                stability = "устойчивый узел"
            else:
                stability = "устойчивый вырожденный узел"
        elif det > 0 and tr > 0:
            if discriminant < 0:
                stability = "неустойчивый фокус"
            elif discriminant > 0:
                stability = "неустойчивый узел"
            else:
                stability = "неустойчивый вырожденный узел"
        elif det < 0:
            stability = "седло"
        elif tr == 0 and det > 0:
            stability = "центр"
        else:
            stability = "нейтральный (возможна бифуркация)"

        return stability, tr, det, eigvals, discriminant

    def analyze_stability_3d(self, point):
        """Анализ устойчивости для 3×3 системы"""
        a, b, c = point

        # Параметры
        pa = self.params['pa']
        ka = self.params['ka']
        # ga не используется в Якобиане, но сохраняем для полноты
        pb = self.params['pb']
        kb = self.params['kb']
        # gb не используется в Якобиане
        pc = self.params['pc']

        # Матрица Якоби 3×3
        J = np.zeros((3, 3))

        # Производные для первого уравнения (a)
        denom1 = 1 + ka * b ** 2
        J[0, 0] = -pa  # df₁/da
        J[0, 1] = -pa * c * (2 * ka * b) / (denom1 ** 2)  # df₁/db
        J[0, 2] = pa / denom1  # df₁/dc

        # Производные для второго уравнения (b)
        denom2 = 1 + kb * a ** 2 * c
        J[1, 0] = -pb * (2 * kb * a * c) / (denom2 ** 2)  # df₂/da
        J[1, 1] = -pb  # df₂/db
        J[1, 2] = -pb * (kb * a ** 2) / (denom2 ** 2)  # df₂/dc

        # Производные для третьего уравнения (c)
        J[2, 0] = -pc * c  # df₃/da
        J[2, 1] = pc  # df₃/db
        J[2, 2] = -pc * a  # df₃/dc

        # Анализ
        eigvals = np.linalg.eigvals(J)
        tr = np.trace(J)
        det = np.linalg.det(J)

        # Определяем тип устойчивости для 3D
        real_parts = np.real(eigvals)

        if all(rp < -1e-10 for rp in real_parts):
            if any(np.iscomplex(eig) for eig in eigvals):
                stability = "устойчивый фокус"
            else:
                stability = "устойчивый узел (3D)"
        elif any(rp > 1e-10 for rp in real_parts):
            if all(rp > 1e-10 for rp in real_parts):
                if any(np.iscomplex(eig) for eig in eigvals):
                    stability = "неустойчивый фокус"
                else:
                    stability = "неустойчивый узел"
            else:
                # Часть собственных значений положительные, часть отрицательные
                if sum(1 for rp in real_parts if rp > 1e-10) == 1:
                    stability = "седло (1 неустойчивое направление)"
                else:
                    stability = "седло (2 неустойчивых направления)"
        elif any(abs(rp) < 1e-10 for rp in real_parts):
            stability = "бифуркация (нейтральный)"
        else:
            stability = "граничный случай"

        return stability, tr, det, eigvals

    # ====================== МЕТОДЫ ИНТЕРФЕЙСА ======================

    def validate_parameters(self):
        """Проверка корректности введенных параметров"""
        for param, entry in self.param_entries.items():
            value_str = entry.get().strip()

            if not value_str:
                return False, f"Параметр {param} не может быть пустым"

            try:
                value = float(value_str)
            except ValueError:
                return False, f"Параметр {param} должен быть числом"

            if value < 0:
                return False, f"Параметр {param} должен быть неотрицательным"

            # Параметры, которые не могут быть нулевыми
            non_zero_params = ['pa', 'pb', 'pc', 'Da', 'Db', 'Dc','ga', 'gb']
            if param in non_zero_params and value == 0:
                return False, f"Параметр {param} не может быть нулевым"

        return True, "Все параметры корректны"

    def apply_parameters(self):
        """Применение параметров из полей ввода"""
        valid, message = self.validate_parameters()

        if not valid:
            messagebox.showerror("Ошибка", message)
            return

        # Обновляем параметры
        for param, entry in self.param_entries.items():
            self.params[param] = float(entry.get())

        self.status_var.set("Параметры применены")
        self.log_calculation("=" * 47)
        self.log_calculation("ПАРАМЕТРЫ МОДЕЛИ ОБНОВЛЕНЫ")
        self.log_calculation("=" * 47)
        self.log_calculation(f"ρ_a = {self.params['pa']:.16f}")
        self.log_calculation(f"ρ_b = {self.params['pb']:.16f}")
        self.log_calculation(f"ρ_c = {self.params['pc']:.16f}")
        self.log_calculation(f"k_a = {self.params['ka']:.16f}")
        self.log_calculation(f"k_b = {self.params['kb']:.16f}")
        self.log_calculation(f"σ_a = {self.params['ga']:.16f}")
        self.log_calculation(f"σ_b = {self.params['gb']:.16f}")
        self.log_calculation(f"D_a = {self.params['Da']:.16f}")
        self.log_calculation(f"D_b = {self.params['Db']:.16f}")
        self.log_calculation(f"D_c = {self.params['Dc']:.16f}")
        self.log_calculation("")
        self.show_parameters_applied_message()

    def log_calculation(self, message):
        """Добавляет сообщение в информационное окно"""
        self.calc_text.insert(tk.END, message + "\n")
        self.calc_text.see(tk.END)
        self.root.update()

    # ====================== НОВЫЕ МЕТОДЫ ДЛЯ КНОПОК ======================

    def plot_equilibrium_points(self):
        """Построение графика функций c₁(a) и c₂(a) с точками равновесия (с центрированием на точки)"""
        try:
            self.status_var.set("Поиск точек равновесия...")
            self.log_calculation("=" * 45)
            self.log_calculation("ТОЧКИ РАВНОВЕСИЯ")
            self.log_calculation("=" * 45)

            # Получаем параметры из текущих настроек
            p_a = self.params['pa']
            k_a = self.params['ka']
            sigma_a = self.params['ga']
            p_b = self.params['pb']
            k_b = self.params['kb']
            sigma_b = self.params['gb']
            p_c = self.params['pc']

            # ========== ВНУТРЕННИЕ ФУНКЦИИ ДЛЯ РАСЧЕТА ==========
            def calculate_c1(a, p_a, k_a, sigma_a):
                """Вычисляет значения c(a) для первой функции"""
                denominator = 2 * (p_a * k_a * a ** 3 - k_a * a ** 2 * sigma_a)

                if abs(denominator) < 1e-10:
                    return np.nan, np.nan

                D1 = p_a ** 2 - 4 * (p_a * k_a * a ** 3 - k_a * a ** 2 * sigma_a) * (a * p_a - sigma_a)

                if D1 < -1e-10:
                    return np.nan, np.nan
                elif abs(D1) < 1e-10:
                    sqrt_D1 = 0
                else:
                    sqrt_D1 = np.sqrt(D1)

                try:
                    c_plus = (p_a + sqrt_D1) / denominator
                    c_minus = (p_a - sqrt_D1) / denominator

                    if np.isinf(c_plus) or np.isinf(c_minus):
                        return np.nan, np.nan

                    c_plus = c_plus if c_plus > 1e-10 else np.nan
                    c_minus = c_minus if c_minus > 1e-10 else np.nan

                    return c_plus, c_minus
                except (ZeroDivisionError, FloatingPointError):
                    return np.nan, np.nan

            def calculate_c2(a, p_b, k_b, sigma_b):
                """Вычисляет значения c(a) для второй функции"""
                denominator = 2 * k_b * a ** 3 * p_b

                if abs(denominator) < 1e-10:
                    return np.nan, np.nan

                term = p_b * a - a ** 2 * k_b * sigma_b
                D2 = term ** 2 + 4 * k_b * a ** 3 * p_b * (p_b + sigma_b)

                if D2 < -1e-10:
                    return np.nan, np.nan
                elif abs(D2) < 1e-10:
                    sqrt_D2 = 0
                else:
                    sqrt_D2 = np.sqrt(D2)

                try:
                    c_plus = (a ** 2 * k_b * sigma_b - p_b * a + sqrt_D2) / denominator
                    c_minus = (a ** 2 * k_b * sigma_b - p_b * a - sqrt_D2) / denominator

                    if np.isinf(c_plus) or np.isinf(c_minus):
                        return np.nan, np.nan

                    c_plus = c_plus if c_plus > 1e-10 else np.nan
                    c_minus = c_minus if c_minus > 1e-10 else np.nan

                    return c_plus, c_minus
                except (ZeroDivisionError, FloatingPointError):
                    return np.nan, np.nan

            # ========== СНАЧАЛА НАЙДЕМ ТОЧКИ ПЕРЕСЕЧЕНИЯ ==========
            # Используем относительно широкий диапазон для поиска
            a_search = np.linspace(0.01, 10.0, 5000)

            c1_plus_search = []
            c1_minus_search = []
            c2_plus_search = []
            c2_minus_search = []

            for a in a_search:
                c1p, c1m = calculate_c1(a, p_a, k_a, sigma_a)
                c2p, c2m = calculate_c2(a, p_b, k_b, sigma_b)
                c1_plus_search.append(c1p)
                c1_minus_search.append(c1m)
                c2_plus_search.append(c2p)
                c2_minus_search.append(c2m)

            c1_plus_search = np.array(c1_plus_search)
            c1_minus_search = np.array(c1_minus_search)
            c2_plus_search = np.array(c2_plus_search)
            c2_minus_search = np.array(c2_minus_search)

            # Поиск пересечений
            all_intersections = []
            for i in range(len(a_search) - 1):
                a1, a2 = a_search[i], a_search[i + 1]

                for c1_branch in [c1_plus_search, c1_minus_search]:
                    for c2_branch in [c2_plus_search, c2_minus_search]:
                        c1_1, c1_2 = c1_branch[i], c1_branch[i + 1]
                        c2_1, c2_2 = c2_branch[i], c2_branch[i + 1]

                        if np.isnan(c1_1) or np.isnan(c1_2) or np.isnan(c2_1) or np.isnan(c2_2):
                            continue

                        diff1 = c1_1 - c2_1
                        diff2 = c1_2 - c2_2

                        if diff1 * diff2 <= 0:
                            if abs(diff2 - diff1) > 1e-15:
                                t = -diff1 / (diff2 - diff1)
                                a_interp = a1 + t * (a2 - a1)
                                c_interp = c1_1 + t * (c1_2 - c1_1)

                                if a_interp > 0 and c_interp > 1e-10:
                                    all_intersections.append((a_interp, c_interp))

            # Удаление дубликатов
            unique_intersections = []
            tol = 1e-6
            for a, c in all_intersections:
                if not any(abs(a - ua) < tol and abs(c - uc) < tol for ua, uc in unique_intersections):
                    unique_intersections.append((a, c))
            unique_intersections = sorted(unique_intersections, key=lambda x: x[0])

            # Сохранение точек
            if unique_intersections:
                self.graphic_equilibrium_points = unique_intersections.copy()

            else:
                self.graphic_equilibrium_points = []
                self.log_calculation("\nВНИМАНИЕ: Графические точки не найдены!")

            # ========== ОПРЕДЕЛЕНИЕ ДИАПАЗОНОВ ДЛЯ ГРАФИКА (ЦЕНТРИРОВАНИЕ) ==========
            if unique_intersections:
                # Находим минимальное и максимальное a среди точек
                a_points = [a for a, _ in unique_intersections]
                c_points = [c for _, c in unique_intersections]

                a_min_point = min(a_points)
                a_max_point = max(a_points)
                c_min_point = min(c_points)
                c_max_point = max(c_points)

                # Определяем поля (отступы) - 20% от диапазона, но не меньше 0.5
                a_range = a_max_point - a_min_point
                c_range = c_max_point - c_min_point

                if a_range < 0.1:  # Если точки очень близко
                    a_padding = 3
                else:
                    a_padding = max(a_range * 0.3, 0.3)

                if c_range < 0.1:
                    c_padding = 10
                else:
                    c_padding = max(c_range * 0.3, 0.3)

                # Устанавливаем пределы
                a_min_plot = max(0, a_min_point - a_padding)
                a_max_plot = a_max_point + a_padding

                c_min_plot = max(0, c_min_point - c_padding)
                c_max_plot = c_max_point + c_padding
            else:
                # Если точек нет, используем стандартный диапазон
                a_min_plot = 0
                a_max_plot = 5.0
                c_min_plot = 0
                c_max_plot = 10.0

            # ========== СОЗДАЕМ СЕТКУ ДЛЯ ПОСТРОЕНИЯ (ТОЛЬКО В НУЖНОМ ДИАПАЗОНЕ) ==========
            # Расширяем диапазон для построения графиков (чтобы видеть подход к точкам)
            a_plot_min = max(0, a_min_plot - 0.5)
            a_plot_max = a_max_plot + 0.5
            c_plot_max = c_max_plot + 1.0

            # Создаем детальную сетку в нужном диапазоне
            a_values = np.concatenate([
                np.linspace(a_plot_min, a_plot_max, 5000)
            ])

            # Вычисление значений для построения
            c1_plus_values = []
            c1_minus_values = []
            c2_plus_values = []
            c2_minus_values = []

            for a in a_values:
                c1_plus, c1_minus = calculate_c1(a, p_a, k_a, sigma_a)
                c2_plus, c2_minus = calculate_c2(a, p_b, k_b, sigma_b)

                c1_plus_values.append(c1_plus)
                c1_minus_values.append(c1_minus)
                c2_plus_values.append(c2_plus)
                c2_minus_values.append(c2_minus)

            c1_plus_values = np.array(c1_plus_values)
            c1_minus_values = np.array(c1_minus_values)
            c2_plus_values = np.array(c2_plus_values)
            c2_minus_values = np.array(c2_minus_values)

            # Функция для нахождения непрерывных участков
            def find_valid_range(a_vals, c_vals):
                valid_indices = ~np.isnan(c_vals)
                if not np.any(valid_indices):
                    return []

                ranges = []
                start = None
                for i, valid in enumerate(valid_indices):
                    if valid and start is None:
                        start = i
                    elif not valid and start is not None:
                        if i - start > 1:
                            ranges.append((start, i))
                        start = None
                if start is not None and len(valid_indices) - start > 1:
                    ranges.append((start, len(valid_indices)))
                return ranges

            # Построение графика
            self.fig.clear()
            ax = self.fig.add_subplot(111)

            # Построение ветвей первой функции (синим)
            c1_plus_ranges = find_valid_range(a_values, c1_plus_values)
            c1_minus_ranges = find_valid_range(a_values, c1_minus_values)

            for start, end in c1_plus_ranges:
                # Обрезаем по диапазону c для красоты
                y_vals = c1_plus_values[start:end]
                mask = y_vals <= c_plot_max * 1.5
                if np.any(mask):
                    ax.plot(a_values[start:end][mask], y_vals[mask],
                            'b-', linewidth=2, alpha=0.7,
                            label='c₁(a)' if start == c1_plus_ranges[0][0] else "")

            for start, end in c1_minus_ranges:
                y_vals = c1_minus_values[start:end]
                mask = y_vals <= c_plot_max * 1.5
                if np.any(mask):
                    ax.plot(a_values[start:end][mask], y_vals[mask],
                            'b-', linewidth=2, alpha=0.7, label='_nolegend_')

            # Построение ветвей второй функции (красным)
            c2_plus_ranges = find_valid_range(a_values, c2_plus_values)
            c2_minus_ranges = find_valid_range(a_values, c2_minus_values)

            for start, end in c2_plus_ranges:
                y_vals = c2_plus_values[start:end]
                mask = y_vals <= c_plot_max * 1.5
                if np.any(mask):
                    ax.plot(a_values[start:end][mask], y_vals[mask],
                            'r-', linewidth=2, alpha=0.7,
                            label='c₂(a)' if start == c2_plus_ranges[0][0] else "")

            for start, end in c2_minus_ranges:
                y_vals = c2_minus_values[start:end]
                mask = y_vals <= c_plot_max * 1.5
                if np.any(mask):
                    ax.plot(a_values[start:end][mask], y_vals[mask],
                            'r-', linewidth=2, alpha=0.7, label='_nolegend_')

            # Отметка точек пересечения
            if unique_intersections:
                intersection_a = [a for a, _ in unique_intersections]
                intersection_c = [c for _, c in unique_intersections]
                ax.plot(intersection_a, intersection_c, 'go', markersize=10,
                        label='Точки равновесия', zorder=5)

                # Добавляем подписи координат точек
                for a, c in unique_intersections:
                    ax.annotate(f'({a:.3f}, {c:.3f})',
                                xy=(a, c), xytext=(5, 5),
                                textcoords='offset points',
                                fontsize=8, color='darkgreen',
                                bbox=dict(boxstyle='round,pad=0.2',
                                          facecolor='white', alpha=0.7))

            # Настройка графика
            ax.set_xlabel('a', fontsize=12)
            ax.set_ylabel('c(a)', fontsize=12)
            ax.set_title(f'Графики функций c₁(a) и c₂(a) с точками равновесия\n'
                         f'ρ_a={p_a}, k_a={k_a}, σ_a={sigma_a}, ρ_b={p_b}, k_b={k_b}, σ_b={sigma_b}',
                         fontsize=11)
            ax.grid(True, alpha=0.3)
            ax.legend(loc='best', fontsize=10)
            ax.axhline(y=0, color='gray', linestyle='--', linewidth=1, alpha=0.5)
            ax.axvline(x=0, color='gray', linestyle='--', linewidth=1, alpha=0.5)

            # Установка пределов (центрированных на точки)
            ax.set_xlim(a_plot_min, a_plot_max)
            ax.set_ylim(0, c_plot_max)

            self.fig.tight_layout()
            self.canvas.draw()

            # Логирование результатов
            self.log_calculation(f"\nПараметры: ρ_a={p_a}, k_a={k_a}, σ_a={sigma_a}, "
                                 f"ρ_b={p_b}, k_b={k_b}, σ_b={sigma_b}")
            self.log_calculation(f"Найдено точек равновесия: {len(unique_intersections)}")

            for i, (a, c) in enumerate(unique_intersections, 1):
                b = a * c
                self.log_calculation(f"Точка {i}: a={a:.8f}, b={b:.8f}, c={c:.8f}")

            self.status_var.set("Точки равновесия построены")

        except Exception as e:
            self.log_calculation(f"Ошибка: {str(e)}")
            import traceback
            self.log_calculation(traceback.format_exc())
            messagebox.showerror("Ошибка", f"Ошибка при построении точек равновесия: {str(e)}")
            self.status_var.set("Ошибка построения")

    def plot_neumann_table(self):
        """Метод Ньютона - уточнение точек равновесия"""
        try:
            self.log_calculation("=" * 45)
            self.log_calculation("МЕТОД НЬЮТОНА")
            self.log_calculation("=" * 45)

            # Получаем параметры
            p_a = self.params['pa']
            k_a = self.params['ka']
            sigma_a = self.params['ga']
            p_b = self.params['pb']
            k_b = self.params['kb']
            sigma_b = self.params['gb']
            p_c = self.params['pc']

            self.log_calculation(f"Параметры модели:")
            self.log_calculation(f"  ρ_a = {p_a:.16f}")
            self.log_calculation(f"  k_a = {k_a:.16f}")
            self.log_calculation(f"  σ_a = {sigma_a:.16f}")
            self.log_calculation(f"  ρ_b = {p_b:.16f}")
            self.log_calculation(f"  k_b = {k_b:.16f}")
            self.log_calculation(f"  σ_b = {sigma_b:.16f}")
            self.log_calculation(f"  ρ_c = {p_c:.16f}")
            self.log_calculation("-" * 78)

            # ========== СИСТЕМА УРАВНЕНИЙ 3×3 ==========
            def system(vars):
                """Три уравнения, три переменные: a, b, c"""
                a, b, c = vars
                eq1 = p_a * (c / (1 + k_a * b ** 2) - a) + sigma_a
                eq2 = p_b * (1 / (1 + k_b * a ** 2 * c) - b) + sigma_b
                eq3 = p_c * (b - c * a)
                return np.array([eq1, eq2, eq3])

            # ========== ЯКОБИАН 3×3 ==========
            def jacobian(vars):
                """Матрица Якоби 3×3"""
                a, b, c = vars
                J = np.zeros((3, 3))

                denom1 = 1 + k_a * b ** 2
                J[0, 0] = -p_a
                J[0, 1] = -p_a * c * (2 * k_a * b) / (denom1 ** 2)
                J[0, 2] = p_a / denom1

                denom2 = 1 + k_b * a ** 2 * c
                J[1, 0] = -p_b * (2 * k_b * a * c) / (denom2 ** 2)
                J[1, 1] = -p_b
                J[1, 2] = -p_b * (k_b * a ** 2) / (denom2 ** 2)

                J[2, 0] = -p_c * c
                J[2, 1] = p_c
                J[2, 2] = -p_c * a

                return J

            # ========== МЕТОД НЬЮТОНА ДЛЯ 3×3 ==========
            def newton_refine_3d(initial_point, max_iter=50, tol=1e-15):
                a_init, b_init, c_init = initial_point
                x = np.array([a_init, b_init, c_init], dtype=float)

                self.log_calculation(f"\nНачальное приближение:")
                self.log_calculation(f"  a = {a_init:.16f}")
                self.log_calculation(f"  b = {b_init:.16f}")
                self.log_calculation(f"  c = {c_init:.16f}")

                for i in range(max_iter):
                    F = system(x)
                    J = jacobian(x)
                    residual_norm = np.linalg.norm(F)

                    self.log_calculation(f"\nШаг {i + 1}:")
                    self.log_calculation(f"  a = {x[0]:.16f}, b = {x[1]:.16f}, c = {x[2]:.16f}")
                    self.log_calculation(f"  Невязка = {residual_norm:.6e}")

                    if residual_norm < tol:
                        self.log_calculation(f"\nСХОДИМОСТЬ ДОСТИГНУТА на шаге {i + 1}!")
                        return x, True, i + 1

                    try:
                        delta = np.linalg.solve(J, -F)
                        self.log_calculation(f"  Δa = {delta[0]:.6e}, Δb = {delta[1]:.6e}, Δc = {delta[2]:.6e}")
                    except np.linalg.LinAlgError:
                        self.log_calculation(f"  ОШИБКА: Матрица Якоби вырождена")
                        return x, False, i + 1

                    x = x + delta

                self.log_calculation(f"\nДОСТИГНУТО МАКСИМАЛЬНОЕ ЧИСЛО ИТЕРАЦИЙ ({max_iter})")
                return x, False, max_iter

            # ========== ПОЛУЧАЕМ ТОЧКИ ИЗ СОХРАНЕННЫХ ГРАФИЧЕСКИХ ТОЧЕК ==========
            graphic_points_3d = []

            # Сначала пробуем использовать сохраненные графические точки
            if hasattr(self, 'graphic_equilibrium_points') and self.graphic_equilibrium_points:
                self.log_calculation(
                    f"\nИспользую сохраненные графические точки ({len(self.graphic_equilibrium_points)} шт.)")
                for a, c in self.graphic_equilibrium_points:
                    b = a * c
                    graphic_points_3d.append((a, b, c))
            else:
                # Если сохраненных нет, пробуем найти заново
                self.log_calculation("\nСохраненных графических точек нет. Выполняю графический поиск...")

                def calculate_c1(a):
                    denominator = 2 * (p_a * k_a * a ** 3 - k_a * a ** 2 * sigma_a)
                    if abs(denominator) < 1e-10:
                        return np.nan, np.nan
                    D1 = p_a ** 2 - 4 * (p_a * k_a * a ** 3 - k_a * a ** 2 * sigma_a) * (a * p_a - sigma_a)
                    if D1 < -1e-10:
                        return np.nan, np.nan
                    sqrt_D1 = np.sqrt(max(0, D1))
                    try:
                        c_plus = (p_a + sqrt_D1) / denominator
                        c_minus = (p_a - sqrt_D1) / denominator
                        c_plus = c_plus if c_plus > 1e-10 else np.nan
                        c_minus = c_minus if c_minus > 1e-10 else np.nan
                        return c_plus, c_minus
                    except:
                        return np.nan, np.nan

                def calculate_c2(a):
                    denominator = 2 * k_b * a ** 3 * p_b
                    if abs(denominator) < 1e-10:
                        return np.nan, np.nan
                    term = p_b * a - a ** 2 * k_b * sigma_b
                    D2 = term ** 2 + 4 * k_b * a ** 3 * p_b * (p_b + sigma_b)
                    if D2 < -1e-10:
                        return np.nan, np.nan
                    sqrt_D2 = np.sqrt(max(0, D2))
                    try:
                        c_plus = (a ** 2 * k_b * sigma_b - p_b * a + sqrt_D2) / denominator
                        c_minus = (a ** 2 * k_b * sigma_b - p_b * a - sqrt_D2) / denominator
                        c_plus = c_plus if c_plus > 1e-10 else np.nan
                        c_minus = c_minus if c_minus > 1e-10 else np.nan
                        return c_plus, c_minus
                    except:
                        return np.nan, np.nan

                a_values = np.linspace(0.01, 5.0, 1000)
                graphic_points_2d = []

                for i in range(len(a_values) - 1):
                    a1, a2 = a_values[i], a_values[i + 1]
                    c1_plus1, c1_minus1 = calculate_c1(a1)
                    c1_plus2, c1_minus2 = calculate_c1(a2)
                    c2_plus1, c2_minus1 = calculate_c2(a1)
                    c2_plus2, c2_minus2 = calculate_c2(a2)

                    branches_c1 = [(c1_plus1, c1_plus2), (c1_minus1, c1_minus2)]
                    branches_c2 = [(c2_plus1, c2_plus2), (c2_minus1, c2_minus2)]

                    for (c1_1, c1_2) in branches_c1:
                        for (c2_1, c2_2) in branches_c2:
                            if not (np.isnan(c1_1) or np.isnan(c1_2) or np.isnan(c2_1) or np.isnan(c2_2)):
                                if (c1_1 - c2_1) * (c1_2 - c2_2) <= 0:
                                    a_mid = (a1 + a2) / 2
                                    c_mid = (c1_1 + c2_1) / 2
                                    graphic_points_2d.append((a_mid, c_mid))

                # Удаляем дубликаты
                unique_graphic_points_2d = []
                tol_graph = 1e-6
                for a, c in graphic_points_2d:
                    if not any(abs(a - ua) < tol_graph and abs(c - uc) < tol_graph
                               for ua, uc in unique_graphic_points_2d):
                        unique_graphic_points_2d.append((a, c))

                unique_graphic_points_2d.sort(key=lambda x: x[0])

                for a, c in unique_graphic_points_2d:
                    b = a * c
                    graphic_points_3d.append((a, b, c))

            if not graphic_points_3d:
                self.log_calculation("\nНет графических точек для уточнения.")
                self.log_calculation("Пожалуйста, сначала нажмите 'Точки равновесия' для построения графика.")
                self.fig.clear()
                ax = self.fig.add_subplot(111)
                ax.text(0.5, 0.5, 'Сначала постройте график\n(кнопка "Точки равновесия")',
                        ha='center', va='center', transform=ax.transAxes, fontsize=14)
                self.canvas.draw()
                self.status_var.set("Нет точек для уточнения")
                return

            self.log_calculation(f"\nРезультаты графического поиска:")
            self.log_calculation(f"  Найдено точек: {len(graphic_points_3d)}")

            for i, (a, b, c) in enumerate(graphic_points_3d, 1):
                self.log_calculation(f"  Точка {i}: a={a:.10f}, b={b:.10f}, c={c:.10f}")

            # ========== УТОЧНЕНИЕ МЕТОДОМ НЬЮТОНА ==========
            self.log_calculation("\n" + "-" * 78)
            self.log_calculation("МЕТОД НЬЮТОНА")
            self.log_calculation("-" * 78)

            refined_points_3d = []

            for idx, (a_init, b_init, c_init) in enumerate(graphic_points_3d, 1):
                self.log_calculation(f"\nУТОЧНЕНИЕ ТОЧКИ №{idx}")

                refined, converged, iterations = newton_refine_3d((a_init, b_init, c_init))

                if converged:
                    unique = True
                    for existing in refined_points_3d:
                        if np.linalg.norm(existing - refined) < 1e-10:
                            unique = False
                            break

                    if unique and all(x > 0 for x in refined):
                        refined_points_3d.append(refined)
                        self.log_calculation(f"\nТОЧКА {idx} УСПЕШНО УТОЧНЕНА!")
                    else:
                        self.log_calculation(f"\nТочка {idx} не уникальна или содержит отрицательные значения")
                else:
                    self.log_calculation(f"\nУточнение точки {idx} не удалось")

            refined_points_3d.sort(key=lambda x: (x[0], x[1], x[2]))

            # ========== СОХРАНЯЕМ УТОЧНЕННЫЕ ТОЧКИ ==========
            if refined_points_3d:
                self.refined_equilibrium_points = refined_points_3d.copy()
                self.equilibrium_points_found = True
                self.log_calculation("\n" + "-" * 78)
                self.log_calculation("Уточненные точки равновесия СОХРАНЕНЫ для последующего использования")
                self.log_calculation(f"Всего сохранено точек: {len(refined_points_3d)}")
                self.log_calculation("-" * 78)
            else:
                self.equilibrium_points_found = False

            # ========== АНАЛИЗ УСТОЙЧИВОСТИ (3×3) ==========
            self.log_calculation("\n" + "-" * 78)
            self.log_calculation("АНАЛИЗ УСТОЙЧИВОСТИ НАЙДЕННЫХ ТОЧЕК")
            self.log_calculation("-" * 78)

            # Создаем таблицу для отображения
            self.fig.clear()
            ax = self.fig.add_subplot(111)
            ax.axis('tight')
            ax.axis('off')

            if not refined_points_3d:
                ax.text(0.5, 0.5, 'Не удалось уточнить точки равновесия',
                        ha='center', va='center', transform=ax.transAxes, fontsize=14)
                self.canvas.draw()
                self.status_var.set("Уточнение не удалось")
                return

            table_data = []
            for i, (a, b, c) in enumerate(refined_points_3d, 1):
                # Используем analyze_stability_3d для анализа
                stability, tr, det, eigvals = self.analyze_stability_3d((a, b, c))

                # Формируем текст устойчивости для таблицы
                if "устойчивый" in stability:
                    stability_text = "Устойчива"
                elif "неустойчивый" in stability:
                    stability_text = "Неустойчива"
                else:
                    stability_text = stability

                table_data.append([f"{i}", f"{a:.16f}", f"{b:.16f}", f"{c:.16f}", stability_text])

                # Подробный вывод в историю
                self.log_calculation(f"\nУТОЧНЕННАЯ ТОЧКА {i}:")
                self.log_calculation(f"  Координаты:")
                self.log_calculation(f"    a = {a:.16f}")
                self.log_calculation(f"    b = {b:.16f}")
                self.log_calculation(f"    c = {c:.16f}")
                self.log_calculation(f"  Анализ устойчивости:")
                self.log_calculation(f"    Тип: {stability}")
                self.log_calculation(f"    След матрицы: tr = {tr:.16f}")
                self.log_calculation(f"    Детерминант: det = {det:.16f}")
                self.log_calculation(f"    Собственные значения:")
                for j, eig in enumerate(eigvals, 1):
                    if np.iscomplex(eig):
                        self.log_calculation(f"      λ{j} = {eig.real:.16f} + {eig.imag:.16f}i")
                    else:
                        self.log_calculation(f"      λ{j} = {eig:.16f}")

            # Создаем таблицу для отображения на графике
            columns = ['№', 'a', 'b', 'c', 'Устойчивость']
            table = ax.table(cellText=table_data, colLabels=columns,
                             cellLoc='center', loc='center',
                             colColours=['#e6f2ff'] * 5)

            # Настройка внешнего вида таблицы
            table.auto_set_font_size(False)
            table.set_fontsize(10)
            table.scale(1.0, 1.5)

            # Заголовок
            ax.set_title(f"МЕТОД НЬЮТОНА\n"
                         f"ρ_a={p_a:.4f}, k_a={k_a:.4f}, σ_a={sigma_a:.4f}, "
                         f"ρ_b={p_b:.4f}, k_b={k_b:.4f}, σ_b={sigma_b:.4f}, ρ_c={p_c:.4f}",
                         fontsize=11, pad=10)

            self.canvas.draw()

            # Итоговый вывод
            self.log_calculation("\n" + "-" * 78)
            self.log_calculation("ИТОГОВЫЙ РЕЗУЛЬТАТ")
            self.log_calculation("-" * 78)
            self.log_calculation(f"  Всего точек пересечения графиков: {len(graphic_points_3d)}")
            self.log_calculation(f"  Успешно уточнено: {len(refined_points_3d)}")
            self.log_calculation("-" * 78)

            self.status_var.set("Метод Ньютона выполнен")

        except Exception as e:
            self.log_calculation(f"ОШИБКА: {str(e)}")
            messagebox.showerror("Ошибка", f"Ошибка в методе Ньютона: {str(e)}")
            self.status_var.set("Ошибка выполнения")
            import traceback
            self.log_calculation(traceback.format_exc())

    def get_refined_equilibrium_points(self):
        """
        Возвращает уточненные точки равновесия из метода Ньютона.
        Если точки не найдены, предлагает запустить метод Ньютона.
        """
        if not self.equilibrium_points_found or not self.refined_equilibrium_points:
            self.log_calculation("Уточненные точки равновесия не найдены.")
            response = messagebox.askyesno("Внимание",
                                           "Сначала необходимо выполнить 'Метод Ньютона' для уточнения точек.\n\n"
                                           "Запустить метод Ньютона сейчас?")
            if response:
                self.plot_neumann_table()
            else:
                return []

        return self.refined_equilibrium_points

    def plot_turing_analysis(self):
        """Анализ тьюринговской неустойчивости"""
        try:
            self.log_calculation("=" * 45)
            self.log_calculation("АНАЛИЗ НЕУСТОЙЧИВОСТИ ТЬЮРИНГА")
            self.log_calculation("=" * 45)

            # Получаем уточненные точки из метода Ньютона
            refined_points = self.get_refined_equilibrium_points()

            if not refined_points:
                self.log_calculation("Нет уточненных точек равновесия. Анализ отменен.")
                self.status_var.set("Анализ Тьюринга отменен")
                return

            # Создаем диалог выбора точки (только из уточненных точек)
            choice_dialog = tk.Toplevel(self.root)
            choice_dialog.title("Выбор точки равновесия")
            choice_dialog.geometry("550x450")
            choice_dialog.transient(self.root)
            choice_dialog.grab_set()

            # Центрируем окно
            choice_dialog.update_idletasks()
            x = (choice_dialog.winfo_screenwidth() // 2) - (550 // 2)
            y = (choice_dialog.winfo_screenheight() // 2) - (450 // 2)
            choice_dialog.geometry(f"+{x}+{y}")

            ttk.Label(choice_dialog, text="Выберите точку равновесия:",
                      style='Title.TLabel').pack(pady=10)

            points_list = tk.Listbox(choice_dialog, height=12, font=('Courier', 10))
            points_list.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

            # Получаем параметры для анализа устойчивости
            p_a = self.params['pa']
            k_a = self.params['ka']
            p_b = self.params['pb']
            k_b = self.params['kb']
            p_c = self.params['pc']

            points_info = []
            for i, (a, b, c) in enumerate(refined_points, 1):
                # Вычисляем матрицу Якоби для проверки устойчивости
                denom1 = 1 + k_a * b ** 2
                denom2 = 1 + k_b * a ** 2 * c

                J_test = np.array([
                    [-p_a, -p_a * c * (2 * k_a * b) / (denom1 ** 2), p_a / denom1],
                    [-p_b * (2 * k_b * a * c) / (denom2 ** 2), -p_b, -p_b * k_b * a ** 2 / (denom2 ** 2)],
                    [-p_c * c, p_c, -p_c * a]
                ])
                eigvals = np.linalg.eigvals(J_test)
                max_real = np.max(np.real(eigvals))

                if max_real < -1e-10:
                    stability = "УСТОЙЧИВА"
                elif max_real > 1e-10:
                    stability = "НЕУСТОЙЧИВА"
                else:
                    stability = "ГРАНИЧНАЯ"

                points_info.append((a, b, c, stability))
                points_list.insert(tk.END, f"{i}: a={a:.8f}, b={b:.8f}, c={c:.8f}  [{stability}]")

            selected_index = [None]

            def on_select():
                if points_list.curselection():
                    selected_index[0] = points_list.curselection()[0]
                    choice_dialog.destroy()

            def on_cancel():
                selected_index[0] = None
                choice_dialog.destroy()

            btn_frame = ttk.Frame(choice_dialog)
            btn_frame.pack(pady=10)

            tk.Button(btn_frame, text="Выбрать", command=on_select,
                      bg='#4da6ff', fg='white', width=12, font=('Arial', 10, 'bold')).pack(side=tk.LEFT, padx=10)
            tk.Button(btn_frame, text="Отмена", command=on_cancel,
                      bg='#cccccc', fg='black', width=12, font=('Arial', 10, 'bold')).pack(side=tk.LEFT, padx=10)

            self.root.wait_window(choice_dialog)

            if selected_index[0] is None:
                self.log_calculation("Выбор точки отменен")
                self.status_var.set("Анализ Тьюринга отменен")
                return

            a0, b0, c0, stability = points_info[selected_index[0]]

            self.log_calculation(f"\nВыбрана точка {selected_index[0] + 1}:")
            self.log_calculation(f"  a = {a0:.16f}")
            self.log_calculation(f"  b = {b0:.16f}")
            self.log_calculation(f"  c = {c0:.16f}")
            self.log_calculation(f"  Устойчивость: {stability}")

            # Проверяем устойчивость точки
            if "НЕУСТОЙЧИВА" in stability:
                self.log_calculation("\nВНИМАНИЕ: Точка неустойчива!")
                self.log_calculation("Тьюринговская неустойчивость анализируется только для устойчивых точек.")
                response = messagebox.askyesno("Предупреждение",
                                               "Выбранная точка равновесия НЕУСТОЙЧИВА!\n"
                                               "Тьюринговская неустойчивость возможна только из устойчивого состояния.\n\n"
                                               "Продолжить анализ?")
                if not response:
                    return

            # ========== ДАЛЬШЕ ИДЕТ ВЕСЬ ОСТАЛЬНОЙ КОД АНАЛИЗА ТЬЮРИНГА ==========
            # (СОХРАНЯЕМ КАК БЫЛО - С КНОПКОЙ ВЫБОРА ГАРМОНИКИ, ГРАФИКАМИ И Т.Д.)

            # Выбор гармоники
            mode_dialog = tk.Toplevel(self.root)
            mode_dialog.title("Выбор гармоники")
            mode_dialog.geometry("350x180")
            mode_dialog.transient(self.root)
            mode_dialog.grab_set()

            tk.Label(mode_dialog, text="Введите номер гармоники (0-20):", font=('Arial', 10)).pack(pady=10)
            tk.Label(mode_dialog, text="Для m=0 диффузия не влияет на определитель матрицы"
                                       "",
                     font=('Arial', 8), fg='gray').pack()

            mode_var = tk.StringVar(value="1")
            entry = tk.Entry(mode_dialog, textvariable=mode_var, width=10, font=('Arial', 12))
            entry.pack(pady=5)
            entry.focus()

            mode_result = [None]

            def on_mode_ok():
                try:
                    value = int(mode_var.get())
                    if 0 <= value <= 20:
                        mode_result[0] = value
                        mode_dialog.destroy()
                    else:
                        messagebox.showerror("Ошибка", "Введите число от 0 до 20")
                except ValueError:
                    messagebox.showerror("Ошибка", "Введите целое число")

            def on_mode_cancel():
                mode_result[0] = None
                mode_dialog.destroy()

            btn_frame = ttk.Frame(mode_dialog)
            btn_frame.pack(pady=10)

            tk.Button(btn_frame, text="OK", command=on_mode_ok,
                      bg='#4da6ff', fg='white', width=10).pack(side=tk.LEFT, padx=5)
            tk.Button(btn_frame, text="Отмена", command=on_mode_cancel,
                      bg='#cccccc', fg='black', width=10).pack(side=tk.LEFT, padx=5)

            entry.bind('<Return>', lambda e: on_mode_ok())

            self.root.wait_window(mode_dialog)

            if mode_result[0] is None:
                self.log_calculation("Выбор гармоники отменен")
                self.status_var.set("Анализ Тьюринга отменен")
                return

            m_display = mode_result[0]
            self.log_calculation(f"Выбрана гармоника m = {m_display}")

            # Создаем объект TuringAnalysis
            turing = TuringAnalysis(self.params, (a0, b0, c0))

            # Анализируем моды
            m_max = 20
            L = 60.0
            results = turing.analyze_modes(m_max, L)
            critical = turing.find_critical_mode(m_max, L)

            # Логирование
            self.log_calculation("-" * 78)
            self.log_calculation(
                f"Параметры диффузии: Da={self.params['Da']:.6f}, Db={self.params['Db']:.6f}, Dc={self.params['Dc']:.6f}")
            self.log_calculation("-" * 45)
            self.log_calculation(f"{'№ гармоники':^6} | {'λ_max':^12} | {'статус':^10}")
            self.log_calculation("-" * 45)

            for i, m in enumerate(results['m']):
                status = "НЕУСТОЙЧИВА" if not results['stable'][i] else "устойчива"
                self.log_calculation(f"{m:6d} | {results['lambda_max_real'][i]:+12.8f} | {status:^10}")

            self.log_calculation("-" * 45)

            if critical['critical_mode']:
                m_crit = critical['critical_mode']['m']
                self.log_calculation(f"РЕЗУЛЬТАТ: Обнаружена Тьюринговская неустойчивость")
                self.log_calculation(f"  Критический номер гармоники: m={m_crit}")
                self.log_calculation(f"  Инкремент: λ={critical['critical_mode']['lambda_max']:.6f}")
                self.log_calculation(f"  Неустойчивых мод: {len(critical['unstable_modes'])}")
            else:
                self.log_calculation(f"РЕЗУЛЬТАТ: {critical['message']}")
            self.log_calculation("=" * 45)

            # Строим графики (как было)
            self.fig.clear()

            ax1 = self.fig.add_subplot(121)

            m_vals = results['m'][1:]
            lambda_vals = results['lambda_max_real'][1:]

            ax1.axhline(y=0, color='black', linestyle='-', linewidth=0.5, alpha=0.5)
            ax1.axvline(x=0, color='black', linestyle='-', linewidth=0.5, alpha=0.5)

            ax1.plot(m_vals, lambda_vals, 'b-', linewidth=2, marker='o', markersize=4)

            if critical['critical_mode']:
                m_crit = critical['critical_mode']['m']
                ax1.plot(m_crit, critical['critical_mode']['lambda_max'],
                         'ro', markersize=8, label=f'Критический номер гармоники m={m_crit}')

            ax1.set_xlabel('Номер гармоники', fontsize=11)
            ax1.set_ylabel('max(Re(λ))', fontsize=11)
            ax1.set_title('Изменение скорости роста возмущений', fontsize=12)

            ax1.xaxis.set_major_locator(plt.MaxNLocator(integer=True))
            if len(m_vals) > 0:
                ax1.set_xticks(np.arange(min(m_vals), max(m_vals) + 1, 1.0))

            ax1.grid(True, alpha=0.3, linestyle='--')
            ax1.legend(loc='best', fontsize=9)

            ax2 = self.fig.add_subplot(122)
            ax2.axis('tight')
            ax2.axis('off')

            # Вычисляем матрицу Якоби
            denom1 = 1 + k_a * b0 ** 2
            denom2 = 1 + k_b * a0 ** 2 * c0

            J = np.array([
                [-p_a, -p_a * c0 * (2 * k_a * b0) / (denom1 ** 2), p_a / denom1],
                [-p_b * (2 * k_b * a0 * c0) / (denom2 ** 2), -p_b, -p_b * k_b * a0 ** 2 / (denom2 ** 2)],
                [-p_c * c0, p_c, -p_c * a0]
            ])

            # Для заданной гармоники
            q2 = (m_display * np.pi / L) ** 2
            Da = self.params['Da']
            Db = self.params['Db']
            Dc = self.params['Dc']

            J_diff = J.copy()
            J_diff[0, 0] -= q2 * Da
            J_diff[1, 1] -= q2 * Db
            J_diff[2, 2] -= q2 * Dc

            tr_diff = np.trace(J_diff)
            det_diff = np.linalg.det(J_diff)

            # Главные миноры
            M11 = J_diff[1, 1] * J_diff[2, 2] - J_diff[1, 2] * J_diff[2, 1]
            M22 = J_diff[0, 0] * J_diff[2, 2] - J_diff[0, 2] * J_diff[2, 0]
            M33 = J_diff[0, 0] * J_diff[1, 1] - J_diff[0, 1] * J_diff[1, 0]
            sum_minors = M11 + M22 + M33

            eigvals_m = np.linalg.eigvals(J_diff)

            info_text = (
                "1. СТАЦИОНАРНОЕ СОСТОЯНИЕ\n"
                "━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                f"a = {a0:.8f}\n"
                f"b = {b0:.8f}\n"
                f"c = {c0:.8f}\n\n"

                "2. УСТОЙЧИВОСТЬ БЕЗ ДИФФУЗИИ\n"
                "━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                f"tr(J) = {np.trace(J):+.6f}\n"
                f"det(J) = {np.linalg.det(J):+.6f}\n"
                f"Статус: {stability}\n\n"

                f"3. МАТРИЦА С ДИФФУЗИЕЙ (m={m_display})\n"
                "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                f"q² = {q2:.6f}\n"
                f"tr(J_diff) = {tr_diff:+.6f}\n"
                f"det(J_diff) = {det_diff:+.6f}\n"
                f"Σ главных миноров 2-го порядка = {sum_minors:+.6f}\n\n"
            )

            if critical['critical_mode']:
                m_crit = critical['critical_mode']['m']
                info_text += (
                    "4. БИФУРКАЦИЯ ТЬЮРИНГА\n"
                    "━━━━━━━━━━━━━━━━━━━━━━\n"
                    f"Критический номер гармоники: m* = {m_crit}\n"
                    f"Волновое число: k* = {np.sqrt(critical['critical_mode']['q2']):.6f}\n"
                    f"Максимальный инкремент: λ_max = {critical['critical_mode']['lambda_max']:.6f}\n"
                    f"Неустойчивых мод: {len(critical['unstable_modes'])}\n\n"
                )
            else:
                info_text += (
                    "4. РЕЗУЛЬТАТ\n"
                    "━━━━━━━━━━━━\n"
                    f"{critical['message']}\n\n"
                )

            eig_str = ", ".join([f"{x.real:.6f}{f'+{x.imag:.6f}i' if abs(x.imag) > 1e-12 else ''}" for x in eigvals_m])
            info_text += f"Спектр для m={m_display}:\n [{eig_str}]"

            ax2.text(0.05, 0.98, info_text, transform=ax2.transAxes,
                     fontsize=8, verticalalignment='top', family='monospace',
                     bbox=dict(boxstyle='round', facecolor='#f8f9ff',
                               edgecolor='#0066cc', linewidth=1.5, pad=0.5))

            self.fig.suptitle('АНАЛИЗ НЕУСТОЙЧИВОСТИ ТЬЮРИНГА', fontsize=14)
            self.fig.tight_layout()
            self.canvas.draw()

            self.status_var.set("Анализ неустойчивости Тьюринга завершен")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при анализе Тьюринга: {str(e)}")
            self.status_var.set("Ошибка анализа")
            import traceback
            self.log_calculation(traceback.format_exc())

    def export_to_file(self):
        """Экспорт данных в файл"""
        try:
            # Создаем папку для результатов если ее нет
            results_dir = "results"
            if not os.path.exists(results_dir):
                os.makedirs(results_dir)

            # Генерируем имя файла с timestamp
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"results/Анализ_системы_{timestamp}.txt"

            # Собираем данные для экспорта
            output = StringIO()

            output.write("=" * 80 + "\n")
            output.write("АНАЛИЗ МОДЕЛИ МАЙНХАРДТА\n")
            output.write("=" * 80 + "\n\n")

            output.write(f"Дата анализа: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

            # Параметры системы
            output.write("ПАРАМЕТРЫ МОДЕЛИ:\n")
            output.write("-" * 40 + "\n")
            for param, value in self.params.items():
                output.write(f"{param} = {value:.16f}\n")
            output.write("\n")

            # Находим точки равновесия для экспорта
            solutions = self.find_equilibrium_points()

            output.write("ТОЧКИ РАВНОВЕСИЯ (МЕТОД НЬЮТОНА):\n")
            output.write("-" * 60 + "\n")

            if solutions:
                output.write(f"{'№':<5} {'a':<15} {'c':<15} {'Устойчивость':<20}\n")
                output.write("-" * 60 + "\n")

                for i, sol in enumerate(solutions, 1):
                    a, c = sol
                    stability, tr, det, eigvals, discriminant = self.analyze_stability(sol)

                    if "устойчивый" in stability:
                        stability_text = "Устойчива"
                    elif "неустойчивый" in stability:
                        stability_text = "Неустойчива"
                    elif "седло" in stability:
                        stability_text = "Седло"
                    else:
                        stability_text = stability

                    output.write(f"{i:<5} {a:<25.16f} {c:<25.16f} {stability_text:<20}\n")
            else:
                output.write("Точки равновесия не найдены\n")

            output.write("\n" + "=" * 80 + "\n")
            output.write("ИСТОРИЯ РАСЧЕТОВ:\n")
            output.write("-" * 40 + "\n")

            calc_text = self.calc_text.get(1.0, tk.END).strip()
            if calc_text:
                output.write(calc_text + "\n")
            else:
                output.write("Расчеты не выполнялись\n")

            output.write("\n" + "=" * 80 + "\n")
            output.write("КОНЕЦ ОТЧЕТА\n")
            output.write("=" * 80 + "\n")

            # Записываем в файл
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(output.getvalue())

            self.log_calculation(f"\nДанные экспортированы в файл: {filename}")
            self.status_var.set(f"Данные экспортированы")

            messagebox.showinfo("Успех", f"Данные успешно экспортированы в файл:\n{filename}")

        except Exception as e:
            self.log_calculation(f"Ошибка экспорта: {str(e)}")
            messagebox.showerror("Ошибка", f"Ошибка при экспорте данных: {str(e)}")
            self.status_var.set("Ошибка экспорта")

class TuringAnalysis:
    """Анализ тьюринговской неустойчивости для трехкомпонентной системы"""

    def __init__(self, params, equilibrium_point):
        """
        params: словарь с параметрами
        equilibrium_point: (a0, b0, c0) - точка равновесия
        """
        self.params = params
        self.a0, self.b0, self.c0 = equilibrium_point
        self.J = None
        self.calc_jacobian()

    def calc_jacobian(self):
        """Вычисляет матрицу Якоби 3x3 в точке равновесия"""
        a, b, c = self.a0, self.b0, self.c0

        # Кинетические параметры
        pa = self.params['pa']
        pb = self.params['pb']
        pc = self.params['pc']
        ka = self.params['ka']
        kb = self.params['kb']

        # Производные для первого уравнения (a)
        denom1 = 1 + ka * b ** 2
        J11 = -pa  # df_a/da
        J12 = -2 * pa * ka * a * c / (denom1 ** 2)  # df_a/db (исправлено)
        J13 = pa / denom1  # df_a/dc

        # Производные для второго уравнения (b)
        denom2 = 1 + kb * a ** 2 * c
        J21 = -2 * pb * kb * a * c / (denom2 ** 2)  # df_b/da
        J22 = -pb  # df_b/db
        J23 = -pb * kb * a ** 2 / (denom2 ** 2)  # df_b/dc

        # Производные для третьего уравнения (c)
        J31 = -pc * c  # df_c/da
        J32 = pc  # df_c/db
        J33 = -pc * a  # df_c/dc

        self.J = np.array([[J11, J12, J13],
                           [J21, J22, J23],
                           [J31, J32, J33]])

        return self.J

    def check_homogeneous_stability(self):
        """Проверка устойчивости однородного состояния (m=0)"""
        eigvals = np.linalg.eigvals(self.J)

        # Условия устойчивости для 3x3
        tr = np.trace(self.J)
        det = np.linalg.det(self.J)

        # Главные миноры
        M2 = self.J[0, 0] * self.J[1, 1] - self.J[0, 1] * self.J[1, 0]

        # Проверяем условия Рауса-Гурвица
        is_stable = (tr < 0) and (det < 0) and (tr * M2 > det) and all(np.real(eigvals) < -1e-10)

        return is_stable, eigvals

    def dispersion_relation(self, m, L=1.0):
        """
        Вычисляет собственные числа для моды m
        m - номер моды (0,1,2,...)
        L - характерный размер системы
        """
        q = (m * np.pi / L) ** 2  # k^2

        Da = self.params.get('Da', 0.1)
        Db = self.params.get('Db', 0.1)
        Dc = self.params.get('Dc', 0.1)

        # Матрица с учетом диффузии
        J_diff = self.J.copy()
        J_diff[0, 0] -= q * Da
        J_diff[1, 1] -= q * Db
        J_diff[2, 2] -= q * Dc

        # Собственные числа
        eigvals = np.linalg.eigvals(J_diff)

        return eigvals, J_diff

    def analyze_modes(self, m_max=20, L=1.0):
        """
        Анализирует все моды от 0 до m_max
        Возвращает словарь с результатами
        """
        results = {
            'm': [],
            'q2': [],  # k^2
            'lambda_max_real': [],  # максимальная действительная часть
            'all_lambdas': [],  # все собственные числа
            'stable': []  # устойчива ли мода
        }

        for m in range(m_max + 1):
            eigvals, _ = self.dispersion_relation(m, L)
            q2 = (m * np.pi / L) ** 2

            results['m'].append(m)
            results['q2'].append(q2)
            results['all_lambdas'].append(eigvals)

            max_real = np.max(np.real(eigvals))
            results['lambda_max_real'].append(max_real)
            results['stable'].append(max_real < 0)

        return results

    def find_critical_mode(self, m_max=20, L=1.0):
        """
        Находит критическую моду (первую неустойчивую или с максимальным инкрементом)
        """
        results = self.analyze_modes(m_max, L)

        # Проверяем устойчивость при m=0
        if not results['stable'][0]:
            return {
                'critical_mode': None,
                'message': 'Однородное состояние неустойчиво',
                'results': results,
                'unstable_modes': [],
                'wavelength': None
            }

        # Ищем неустойчивые моды
        unstable_modes = []
        for i in range(1, len(results['m'])):
            if not results['stable'][i]:
                unstable_modes.append({
                    'm': results['m'][i],
                    'q2': results['q2'][i],
                    'lambda_max': results['lambda_max_real'][i]
                })

        if not unstable_modes:
            return {
                'critical_mode': None,
                'message': 'Все гармоники устойчивы',
                'results': results,
                'unstable_modes': [],
                'wavelength': None
            }

        # Берем моду с максимальным инкрементом
        critical = max(unstable_modes, key=lambda x: x['lambda_max'])

        # Характерный масштаб (длина волны)
        wavelength = 2 * np.pi / np.sqrt(critical['q2']) if critical['q2'] > 0 else float('inf')

        return {
            'critical_mode': critical,
            'wavelength': wavelength,
            'unstable_modes': unstable_modes,
            'message': f'Найдена неустойчивая гармоника m={critical["m"]}',
            'results': results
        }


class MainTaskNumericalSolver:
    """Решатель основной задачи (численное решение системы) явной схемой"""

    def __init__(self, params, init_params, grid_params):
        """
        params: словарь с параметрами модели
        init_params: словарь с параметрами начальных условий
        grid_params: словарь с параметрами сетки (n, m, T, L)
        """
        self.params = params.copy()
        self.init_params = init_params.copy()

        # Параметры сетки
        self.n = grid_params['n']
        self.m = grid_params['m']
        self.T = grid_params['T']
        self.L = grid_params['L']

        # Шаги сетки
        self.h = self.L / self.n
        self.tau = self.T / self.m

        # Координаты узлов
        self.x = np.linspace(0, self.L, self.n + 1)
        self.t = np.linspace(0, self.T, self.m + 1)

        # Массивы для хранения решения (a, b, c - как в аналитической части)
        self.a = np.zeros((self.m + 1, self.n + 1))
        self.b = np.zeros((self.m + 1, self.n + 1))
        self.c = np.zeros((self.m + 1, self.n + 1))

        # Статус расчета
        self.calculation_done = False

    def check_stability(self):
        """Проверяет условие устойчивости явной схемы"""
        D_max = max(self.params['Da'], self.params['Db'], self.params['Dc'])
        stability_limit = self.h ** 2 / (2 * D_max) if D_max > 0 else float('inf')
        return self.tau <= stability_limit, stability_limit

    def initial_condition_a(self, x):
        """Начальное условие для a(x,0)"""
        A0 = self.init_params.get('A0_a', 1.0)
        A1 = self.init_params.get('A1_a', 0.0)
        A2 = self.init_params.get('A2_a', 0.0)
        A3 = self.init_params.get('A3_a', 0.0)
        b1 = self.init_params.get('b1', 1)
        b2 = self.init_params.get('b2', 2)
        b3 = self.init_params.get('b3', 3)

        return (A0 + A1 * np.cos(b1 * np.pi * x / self.L) +
                A2 * np.cos(b2 * np.pi * x / self.L) +
                A3 * np.cos(b3 * np.pi * x / self.L))

    def initial_condition_b(self, x):
        """Начальное условие для b(x,0)"""
        A0 = self.init_params.get('A0_b', 1.0)
        A1 = self.init_params.get('A1_b', 0.0)
        A2 = self.init_params.get('A2_b', 0.0)
        A3 = self.init_params.get('A3_b', 0.0)
        b1 = self.init_params.get('b1', 1)
        b2 = self.init_params.get('b2', 2)
        b3 = self.init_params.get('b3', 3)

        return (A0 + A1 * np.cos(b1 * np.pi * x / self.L) +
                A2 * np.cos(b2 * np.pi * x / self.L) +
                A3 * np.cos(b3 * np.pi * x / self.L))

    def initial_condition_c(self, x):
        """Начальное условие для c(x,0)"""
        A0 = self.init_params.get('A0_c', 1.0)
        A1 = self.init_params.get('A1_c', 0.0)
        A2 = self.init_params.get('A2_c', 0.0)
        A3 = self.init_params.get('A3_c', 0.0)
        b1 = self.init_params.get('b1', 1)
        b2 = self.init_params.get('b2', 2)
        b3 = self.init_params.get('b3', 3)

        return (A0 + A1 * np.cos(b1 * np.pi * x / self.L) +
                A2 * np.cos(b2 * np.pi * x / self.L) +
                A3 * np.cos(b3 * np.pi * x / self.L))

    def solve(self, log_callback=None):
        """
        Решает систему уравнений явной разностной схемой
        Уравнения соответствуют аналитической постановке:
        ∂a/∂t = ρ_a (c/(1 + k_a·b²) - a) + σ_a + D_a·Δa
        ∂b/∂t = ρ_b (1/(1 + k_b·a²·c) - b) + σ_b + D_b·Δb
        ∂c/∂t = ρ_c (b - c·a) + D_c·Δc
        """
        try:
            if log_callback:
                log_callback(f"Параметры сетки: n={self.n}, m={self.m}, T={self.T}, L={self.L}")
                log_callback(f"Шаги: h={self.h:.6f}, τ={self.tau:.6f}")

            # Извлекаем параметры для удобства
            pa = self.params['pa']
            pb = self.params['pb']
            pc = self.params['pc']
            ka = self.params['ka']
            kb = self.params['kb']
            ga = self.params['ga']
            gb = self.params['gb']
            Da = self.params['Da']
            Db = self.params['Db']
            Dc = self.params['Dc']

            # Задаем начальные условия (слой j=0)
            for i in range(self.n + 1):
                self.a[0, i] = self.initial_condition_a(self.x[i])
                self.b[0, i] = self.initial_condition_b(self.x[i])
                self.c[0, i] = self.initial_condition_c(self.x[i])

            # Основной цикл по времени
            for j in range(self.m):

                # Внутренние узлы (i = 1, ..., n-1)
                for i in range(1, self.n):
                    # Члены с диффузией (используем центральную разность)
                    diff_a = Da * (self.a[j, i - 1] - 2 * self.a[j, i] + self.a[j, i + 1]) / (self.h ** 2)
                    diff_b = Db * (self.b[j, i - 1] - 2 * self.b[j, i] + self.b[j, i + 1]) / (self.h ** 2)
                    diff_c = Dc * (self.c[j, i - 1] - 2 * self.c[j, i] + self.c[j, i + 1]) / (self.h ** 2)

                    # Кинетические члены (соответствуют аналитической системе)
                    # Для a: ρ_a (c/(1 + k_a·b²) - a) + σ_a
                    kin_a = pa * (self.c[j, i] / (1 + ka * self.b[j, i] ** 2) - self.a[j, i]) + ga

                    # Для b: ρ_b (1/(1 + k_b·a²·c) - b) + σ_b
                    kin_b = pb * (1 / (1 + kb * self.a[j, i] ** 2 * self.c[j, i]) - self.b[j, i]) + gb

                    # Для c: ρ_c (b - c·a)
                    kin_c = pc * (self.b[j, i] - self.c[j, i] * self.a[j, i])

                    # Обновление значений (явная схема)
                    self.a[j + 1, i] = self.a[j, i] + self.tau * (kin_a + diff_a)
                    self.b[j + 1, i] = self.b[j, i] + self.tau * (kin_b + diff_b)
                    self.c[j + 1, i] = self.c[j, i] + self.tau * (kin_c + diff_c)

                # Граничные узлы (условия Неймана: производная на границе = 0)
                # Используем аппроксимацию второго порядка для граничных условий

                # Левый край (i=0): ∂u/∂x = 0 => u(0) = u(1)
                # Для диффузионного члена: ∂²u/∂x² ≈ 2(u(1) - u(0))/h²

                # Для a
                diff_a_left = Da * 2 * (self.a[j, 1] - self.a[j, 0]) / (self.h ** 2)
                kin_a_left = pa * (self.c[j, 0] / (1 + ka * self.b[j, 0] ** 2) - self.a[j, 0]) + ga
                self.a[j + 1, 0] = self.a[j, 0] + self.tau * (kin_a_left + diff_a_left)

                # Для b
                diff_b_left = Db * 2 * (self.b[j, 1] - self.b[j, 0]) / (self.h ** 2)
                kin_b_left = pb * (1 / (1 + kb * self.a[j, 0] ** 2 * self.c[j, 0]) - self.b[j, 0]) + gb
                self.b[j + 1, 0] = self.b[j, 0] + self.tau * (kin_b_left + diff_b_left)

                # Для c
                diff_c_left = Dc * 2 * (self.c[j, 1] - self.c[j, 0]) / (self.h ** 2)
                kin_c_left = pc * (self.b[j, 0] - self.c[j, 0] * self.a[j, 0])
                self.c[j + 1, 0] = self.c[j, 0] + self.tau * (kin_c_left + diff_c_left)

                # Правый край (i=n): ∂u/∂x = 0 => u(n) = u(n-1)
                # Для a
                diff_a_right = Da * 2 * (self.a[j, self.n - 1] - self.a[j, self.n]) / (self.h ** 2)
                kin_a_right = pa * (self.c[j, self.n] / (1 + ka * self.b[j, self.n] ** 2) - self.a[j, self.n]) + ga
                self.a[j + 1, self.n] = self.a[j, self.n] + self.tau * (kin_a_right + diff_a_right)

                # Для b
                diff_b_right = Db * 2 * (self.b[j, self.n - 1] - self.b[j, self.n]) / (self.h ** 2)
                kin_b_right = pb * (1 / (1 + kb * self.a[j, self.n] ** 2 * self.c[j, self.n]) - self.b[j, self.n]) + gb
                self.b[j + 1, self.n] = self.b[j, self.n] + self.tau * (kin_b_right + diff_b_right)

                # Для c
                diff_c_right = Dc * 2 * (self.c[j, self.n - 1] - self.c[j, self.n]) / (self.h ** 2)
                kin_c_right = pc * (self.b[j, self.n] - self.c[j, self.n] * self.a[j, self.n])
                self.c[j + 1, self.n] = self.c[j, self.n] + self.tau * (kin_c_right + diff_c_right)

            self.calculation_done = True
            if log_callback:
                log_callback("Расчет завершен успешно!")

            return True

        except Exception as e:
            if log_callback:
                log_callback(f"ОШИБКА при расчете: {str(e)}")
                import traceback
                log_callback(traceback.format_exc())
            return False

    def get_layer(self, substance, layer_idx):
        """
        Возвращает данные для указанного слоя
        substance: 'a', 'b' или 'c'
        layer_idx: номер слоя
        """
        if layer_idx < 0 or layer_idx > self.m:
            return None, None

        if substance == 'a':
            return self.x, self.a[layer_idx, :]
        elif substance == 'b':
            return self.x, self.b[layer_idx, :]
        elif substance == 'c':
            return self.x, self.c[layer_idx, :]
        else:
            return None, None

    def export_to_file(self, filename, step=1, substances=['a', 'b', 'c']):
        """Экспорт данных в файл"""
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                f.write("РЕЗУЛЬТАТЫ ЧИСЛЕННОГО РЕШЕНИЯ СИСТЕМЫ\n")
                f.write("=" * 80 + "\n")
                f.write(f"Дата: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Параметры сетки: n={self.n}, m={self.m}, T={self.T}, L={self.L}\n")
                f.write(f"Шаги: h={self.h:.6f}, τ={self.tau:.6f}\n\n")

                f.write("ПАРАМЕТРЫ МОДЕЛИ:\n")
                for key, value in self.params.items():
                    f.write(f"  {key} = {value:.6f}\n")
                f.write("\n")

                f.write("ПАРАМЕТРЫ НАЧАЛЬНЫХ УСЛОВИЙ:\n")
                for key, value in self.init_params.items():
                    f.write(f"  {key} = {value}\n")
                f.write("\n")

                # Вывод данных по слоям
                layers = list(range(0, self.m + 1, step))
                f.write(f"Экспортировано слоев: {len(layers)} (шаг={step})\n\n")

                for layer in layers:
                    f.write(f"Слой {layer}: t = {self.t[layer]:.6f}\n")
                    f.write("-" * 70 + "\n")
                    f.write(f"{'x':>10} | {'a(x,t)':>15} | {'b(x,t)':>15} | {'c(x,t)':>15}\n")
                    f.write("-" * 70 + "\n")

                    for i in range(self.n + 1):
                        f.write(f"{self.x[i]:10.6f} | {self.a[layer, i]:15.8f} | "
                                f"{self.b[layer, i]:15.8f} | {self.c[layer, i]:15.8f}\n")
                    f.write("\n")

            return True
        except Exception as e:
            print(f"Ошибка экспорта: {e}")
            return False


class NumericalTab:
    """Вкладка численного решения основной задачи"""

    def __init__(self, parent, main_app):
        self.parent = parent
        self.main_app = main_app
        self.solver = None
        self.setup_tab()

    def setup_tab(self):
        """Создание интерфейса вкладки"""
        # Основной контейнер вкладки
        main_container = ttk.Frame(self.parent, style='Blue.TFrame')
        main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Левая панель - управление
        left_panel = ttk.Frame(main_container, style='Blue.TFrame', width=400)
        left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        left_panel.pack_propagate(False)

        # Правая панель - графики
        right_panel = ttk.Frame(main_container, style='Blue.TFrame')
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.setup_control_panel(left_panel)
        self.setup_plot_area(right_panel)

    def setup_control_panel(self, parent):
        """Левая панель с параметрами (с прокруткой)"""

        # Создаем Canvas с прокруткой для всей левой панели
        canvas = tk.Canvas(parent, bg='#e6f2ff', highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        main_scrollable_frame = ttk.Frame(canvas, style='Blue.TFrame')

        main_scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        # Создаем окно в canvas
        canvas_window = canvas.create_window((0, 0), window=main_scrollable_frame, anchor="nw")

        # Обновляем ширину окна при изменении размера canvas
        def _configure_canvas(event):
            canvas.itemconfig(canvas_window, width=event.width)

        canvas.bind("<Configure>", _configure_canvas)
        canvas.configure(yscrollcommand=scrollbar.set)

        # Привязываем колесо мыши
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # ======== ПАРАМЕТРЫ МОДЕЛИ ========
        params_frame = ttk.LabelFrame(main_scrollable_frame, text="ПАРАМЕТРЫ МОДЕЛИ", style='Blue.TLabelframe')
        params_frame.pack(fill=tk.X, pady=(0, 10), padx=5)

        # Контейнер с двумя колонками
        params_container = ttk.Frame(params_frame, style='Blue.TFrame')
        params_container.pack(fill=tk.X, padx=5, pady=5)

        # Левая колонка
        left_col = ttk.Frame(params_container, style='Blue.TFrame')
        left_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

        # Правая колонка
        right_col = ttk.Frame(params_container, style='Blue.TFrame')
        right_col.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.param_entries = {}

        # Левая колонка: ρ_a, ρ_b, ρ_c, k_a, k_b
        left_params = [
            ('pa', 'ρ_a =', 1.0),
            ('pb', 'ρ_b =', 1.0),
            ('pc', 'ρ_c =', 0.1),
            ('ka', 'k_a =', 5.0),
            ('kb', 'k_b =', 5.0)
        ]

        for param, label_text, default in left_params:
            frame = ttk.Frame(left_col, style='Blue.TFrame')
            frame.pack(fill=tk.X, pady=2)

            ttk.Label(frame, text=label_text, width=8, style='Param.TLabel').pack(side=tk.LEFT)
            entry = ttk.Entry(frame, width=10)
            entry.insert(0, str(default))
            entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
            self.param_entries[param] = entry

        # Правая колонка: σ_a, σ_b, D_a, D_b, D_c
        right_params = [
            ('ga', 'σ_a =', 0.01),
            ('gb', 'σ_b =', 0.01),
            ('Da', 'D_a =', 0.5),
            ('Db', 'D_b =', 0.5),
            ('Dc', 'D_c =', 0.5)
        ]

        for param, label_text, default in right_params:
            frame = ttk.Frame(right_col, style='Blue.TFrame')
            frame.pack(fill=tk.X, pady=2)

            ttk.Label(frame, text=label_text, width=8, style='Param.TLabel').pack(side=tk.LEFT)
            entry = ttk.Entry(frame, width=10)
            entry.insert(0, str(default))
            entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
            self.param_entries[param] = entry

        # ======== НАЧАЛЬНЫЕ УСЛОВИЯ ========
        init_frame = ttk.LabelFrame(main_scrollable_frame, text="НАЧАЛЬНЫЕ УСЛОВИЯ", style='Blue.TLabelframe')
        init_frame.pack(fill=tk.X, pady=(0, 10), padx=5)

        # Простой фрейм для начальных условий
        init_content_frame = ttk.Frame(init_frame, style='Blue.TFrame')
        init_content_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.init_entries = {}

        # Заголовки
        header_frame = ttk.Frame(init_content_frame, style='Blue.TFrame')
        header_frame.pack(fill=tk.X, pady=2)

        ttk.Label(header_frame, text="Параметр", width=6, style='Param.TLabel').grid(row=0, column=0, padx=1)
        ttk.Label(header_frame, text="a(x,0)", width=6, style='Param.TLabel').grid(row=0, column=1, padx=1)
        ttk.Label(header_frame, text="b(x,0)", width=6, style='Param.TLabel').grid(row=0, column=2, padx=1)
        ttk.Label(header_frame, text="c(x,0)", width=6, style='Param.TLabel').grid(row=0, column=3, padx=1)
        ttk.Label(header_frame, text="m₁,m₂,m₃", width=8, style='Param.TLabel').grid(row=0, column=4, padx=1)

        # A0, A1, A2, A3
        init_params = [
            ('A0', 'A₀ =', 0.47),
            ('A1', 'A₁ =', 0.1),
            ('A2', 'A₂ =', 0.0),
            ('A3', 'A₃ =', 0.0)
        ]

        for row, (param_base, label, default) in enumerate(init_params, 1):
            ttk.Label(header_frame, text=label, style='Param.TLabel').grid(row=row, column=0, padx=1, pady=1,
                                                                           sticky=tk.W)

            # Для a
            entry_a = ttk.Entry(header_frame, width=8)
            entry_a.insert(0, str(default))
            entry_a.grid(row=row, column=1, padx=1, pady=1)
            self.init_entries[f'{param_base}_a'] = entry_a

            # Для b
            entry_b = ttk.Entry(header_frame, width=8)
            entry_b.insert(0, str(default))
            entry_b.grid(row=row, column=2, padx=1, pady=1)
            self.init_entries[f'{param_base}_b'] = entry_b

            # Для c
            entry_c = ttk.Entry(header_frame, width=8)
            entry_c.insert(0, str(default))
            entry_c.grid(row=row, column=3, padx=1, pady=1)
            self.init_entries[f'{param_base}_c'] = entry_c

        # Частоты b1, b2, b3 (размещаем справа)
        # b1
        ttk.Label(header_frame, text="m₁ =", style='Param.TLabel').grid(row=1, column=4, padx=1, pady=1, sticky=tk.W)
        entry_b1 = ttk.Entry(header_frame, width=6)
        entry_b1.insert(0, '1')
        entry_b1.grid(row=1, column=4, padx=(20, 1), pady=1, sticky=tk.E)
        self.init_entries['b1'] = entry_b1

        # b2
        ttk.Label(header_frame, text="m₂ =", style='Param.TLabel').grid(row=2, column=4, padx=1, pady=1, sticky=tk.W)
        entry_b2 = ttk.Entry(header_frame, width=6)
        entry_b2.insert(0, '2')
        entry_b2.grid(row=2, column=4, padx=(20, 1), pady=1, sticky=tk.E)
        self.init_entries['b2'] = entry_b2

        # b3
        ttk.Label(header_frame, text="m₃ =", style='Param.TLabel').grid(row=3, column=4, padx=1, pady=1, sticky=tk.W)
        entry_b3 = ttk.Entry(header_frame, width=6)
        entry_b3.insert(0, '3')
        entry_b3.grid(row=3, column=4, padx=(20, 1), pady=1, sticky=tk.E)
        self.init_entries['b3'] = entry_b3

        # ======== ПАРАМЕТРЫ СЕТКИ ========
        grid_frame = ttk.LabelFrame(main_scrollable_frame, text="ПАРАМЕТРЫ РАСЧЕТА", style='Blue.TLabelframe')
        grid_frame.pack(fill=tk.X, pady=(0, 10), padx=5)

        grid_container = ttk.Frame(grid_frame, style='Blue.TFrame')
        grid_container.pack(fill=tk.X, padx=5, pady=5)

        # Параметры сетки в две колонки
        left_grid = ttk.Frame(grid_container, style='Blue.TFrame')
        left_grid.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

        right_grid = ttk.Frame(grid_container, style='Blue.TFrame')
        right_grid.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.grid_entries = {}

        # Левая колонка: n, m
        grid_left_params = [
            ('n', 'Число участков по x (n):', 40),
            ('m', 'Число участков по t (m):', 1000)
        ]

        for param, label_text, default in grid_left_params:
            frame = ttk.Frame(left_grid, style='Blue.TFrame')
            frame.pack(fill=tk.X, pady=2)
            ttk.Label(frame, text=label_text, style='Param.TLabel').pack(anchor=tk.W)
            entry = ttk.Entry(frame, width=12)
            entry.insert(0, str(default))
            entry.pack(fill=tk.X, pady=(0, 5))
            self.grid_entries[param] = entry

        # Правая колонка: L, T
        grid_right_params = [
            ('L', 'Длина отрезка(L):', 60.0),
            ('T', 'Конечное время (T):', 10.0)
        ]

        for param, label_text, default in grid_right_params:
            frame = ttk.Frame(right_grid, style='Blue.TFrame')
            frame.pack(fill=tk.X, pady=2)
            ttk.Label(frame, text=label_text, style='Param.TLabel').pack(anchor=tk.W)
            entry = ttk.Entry(frame, width=12)
            entry.insert(0, str(default))
            entry.pack(fill=tk.X, pady=(0, 5))
            self.grid_entries[param] = entry

        # ======== КНОПКИ УПРАВЛЕНИЯ ========
        buttons_control_frame = ttk.Frame(main_scrollable_frame, style='Blue.TFrame')
        buttons_control_frame.pack(fill=tk.X, pady=10, padx=5)

        # Кнопка "ЗАПУСТИТЬ РАСЧЕТ"
        self.calc_button = tk.Button(buttons_control_frame, text="ЗАПУСТИТЬ РАСЧЕТ",
                                     command=self.start_calculation,
                                     bg='#4da6ff', fg='white',
                                     font=('Arial', 10, 'bold'),
                                     height=1,
                                     relief=tk.RAISED, bd=2)
        self.calc_button.pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)

        # Кнопка "ПРОВЕРКА СХОДИМОСТИ"
        self.convergence_button = tk.Button(buttons_control_frame, text="ПРОВЕРКА СХОДИМОСТИ",
                                            command=self.check_convergence,
                                            bg='#4da6ff', fg='white',
                                            font=('Arial', 10, 'bold'),
                                            height=1,
                                            relief=tk.RAISED, bd=2)
        self.convergence_button.pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)

        # ======== ВИЗУАЛИЗАЦИЯ ========
        viz_frame = ttk.LabelFrame(main_scrollable_frame, text="ВИЗУАЛИЗАЦИЯ", style='Blue.TLabelframe')
        viz_frame.pack(fill=tk.X, pady=(0, 10), padx=5)

        # Выбор вещества
        substance_frame = ttk.Frame(viz_frame, style='Blue.TFrame')
        substance_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(substance_frame, text="Вещество:", style='Param.TLabel').pack(side=tk.LEFT, padx=(0, 5))
        self.substance_var = tk.StringVar(value="a")
        substance_combo = ttk.Combobox(substance_frame, textvariable=self.substance_var,
                                       values=["a", "b", "c"], width=5, state="readonly")
        substance_combo.pack(side=tk.LEFT)

        # Выбор слоев
        layers_frame = ttk.Frame(viz_frame, style='Blue.TFrame')
        layers_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(layers_frame, text="Выбрать слои (через запятую):", style='Param.TLabel').pack(anchor=tk.W)
        self.layers_var = tk.StringVar(value="0,1")
        ttk.Entry(layers_frame, textvariable=self.layers_var).pack(fill=tk.X, pady=5)

        # Кнопки визуализации (2D, 3D и Поверхность в одной строке)
        buttons_frame = ttk.Frame(viz_frame, style='Blue.TFrame')
        buttons_frame.pack(fill=tk.X, padx=5, pady=5)

        self.plot_2d_button = tk.Button(buttons_frame, text="2D слои",
                                        command=self.plot_2d,
                                        bg='#66b3ff', fg='white',
                                        width=10, state=tk.DISABLED)
        self.plot_2d_button.pack(side=tk.LEFT, padx=2, pady=2, expand=True, fill=tk.X)

        self.plot_3d_button = tk.Button(buttons_frame, text="3D слои",
                                        command=self.plot_3d,
                                        bg='#66b3ff', fg='white',
                                        width=10, state=tk.DISABLED)
        self.plot_3d_button.pack(side=tk.LEFT, padx=2, pady=2, expand=True, fill=tk.X)

        # Новая кнопка "Поверхность" (для выбранных слоёв)
        self.plot_surface_selected_button = tk.Button(buttons_frame, text="Поверхность",
                                                      command=self.plot_surface_selected,
                                                      bg='#66b3ff', fg='white',
                                                      width=10, state=tk.DISABLED)
        self.plot_surface_selected_button.pack(side=tk.LEFT, padx=2, pady=2, expand=True, fill=tk.X)

        # Строка: Шаг для поверхности + кнопка "Все слои"
        surface_frame = ttk.Frame(viz_frame, style='Blue.TFrame')
        surface_frame.pack(fill=tk.X, padx=5, pady=5)

        # Верхняя часть - поле ввода и кнопка в одной строке
        top_surface_frame = ttk.Frame(surface_frame, style='Blue.TFrame')
        top_surface_frame.pack(fill=tk.X)

        # Левая часть - шаг для поверхности
        step_frame = ttk.Frame(top_surface_frame, style='Blue.TFrame')
        step_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

        ttk.Label(step_frame, text="Шаг для поверхности:", style='Param.TLabel').pack(side=tk.LEFT, padx=(0, 5))
        self.step_var = tk.StringVar(value="1")
        ttk.Entry(step_frame, textvariable=self.step_var, width=8).pack(side=tk.LEFT)

        # Правая часть - кнопка "Все слои"
        self.plot_surface_all_button = tk.Button(top_surface_frame, text="Все слои",
                                                 command=self.plot_surface,
                                                 bg='#66b3ff', fg='white',
                                                 width=12, state=tk.DISABLED)
        self.plot_surface_all_button.pack(side=tk.RIGHT, padx=2, pady=2)

        # Нижняя часть - поясняющий комментарий (отдельная строка)
        comment_frame = ttk.Frame(surface_frame, style='Blue.TFrame')
        comment_frame.pack(fill=tk.X, pady=(5, 0))

        comment_label = ttk.Label(comment_frame,
                                  text="Шаг 1: каждый слой, Шаг 2: слои 0,2,4... , Шаг 3: 0,3,6...",
                                  style='Param.TLabel',
                                  font=('Arial', 8))
        comment_label.pack(anchor=tk.W)

        # Кнопка экспорта
        self.export_button = tk.Button(viz_frame, text="ВЫВОД ДАННЫХ В ФАЙЛ",
                                       command=self.export_data,
                                       bg='#66b3ff', fg='white',
                                       font=('Arial', 9, 'bold'),
                                       state=tk.DISABLED)
        self.export_button.pack(fill=tk.X, padx=5, pady=10)

        # Информационное окно
        info_frame = ttk.LabelFrame(main_scrollable_frame, text="ИСТОРИЯ РАСЧЕТОВ", style='Blue.TLabelframe')
        info_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5), padx=5)

        # Создаем фрейм для текста и скроллбара
        text_frame = ttk.Frame(info_frame, style='Blue.TFrame')
        text_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Текстовое поле - убираем height, чтобы оно растягивалось
        self.info_text = tk.Text(text_frame,
                                 bg='#f0f8ff',
                                 fg='#0066cc',
                                 font=('Arial', 9),
                                 wrap=tk.WORD)

        # Вертикальный скроллбар
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.info_text.yview)
        self.info_text.configure(yscrollcommand=scrollbar.set)

        # Размещаем с использованием grid для точного контроля
        self.info_text.grid(row=0, column=0, sticky='nsew')
        scrollbar.grid(row=0, column=1, sticky='ns')

        # Настраиваем веса для растягивания
        text_frame.grid_rowconfigure(0, weight=1)
        text_frame.grid_columnconfigure(0, weight=1)

        # Принудительно обновляем
        self.info_text.update_idletasks()

        # Статус бар
        self.status_var = tk.StringVar(value="Готов к работе")
        status_label = tk.Label(main_scrollable_frame, textvariable=self.status_var,
                                bg='#cce5ff', fg='#0066cc', relief=tk.SUNKEN,
                                font=('Arial', 9))
        status_label.pack(fill=tk.X, pady=(5, 5), padx=5)

    def setup_plot_area(self, parent):
        """Правая панель с графиками"""
        plot_frame = ttk.LabelFrame(parent, text="ГРАФИКИ РЕШЕНИЯ", style='Blue.TLabelframe')
        plot_frame.pack(fill=tk.BOTH, expand=True)

        # Контейнер для графика
        graph_container = ttk.Frame(plot_frame, style='Blue.TFrame')
        graph_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Создаем фигуру matplotlib
        self.fig = Figure(figsize=(9, 7), facecolor='white')
        self.canvas = FigureCanvasTkAgg(self.fig, graph_container)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        # Приветственный график
        self.show_welcome_plot()

    def show_welcome_plot(self):
        """Показывает приветственный график"""
        self.fig.clear()
        ax = self.fig.add_subplot(111)
        ax.set_facecolor('#f0f8ff')

        info_text = 'ЧИСЛЕННОЕ РЕШЕНИЕ СИСТЕМЫ\n\n' \
                    '∂a/∂t = ρ_a (c/(1 + k_a·b²) - a) + σ_a + D_a·Δa\n' \
                    '∂b/∂t = ρ_b (1/(1 + k_b·a²·c) - b) + σ_b + D_b·Δb\n' \
                    '∂c/∂t = ρ_c (b - c·a) + D_c·Δc\n\n' \
                    'Для начала работы:\n' \
                    '1. Задайте параметры модели\n' \
                    '2. Задайте начальные условия\n' \
                    '3. Задайте параметры расчета\n' \
                    '4. Нажмите "Запустить расчет"\n' \
                    '5. Используйте кнопки визуализации'

        ax.text(0.5, 0.5, info_text, horizontalalignment='center',
                verticalalignment='center', transform=ax.transAxes,
                fontsize=12, color='#0066cc')
        ax.set_xticks([])
        ax.set_yticks([])
        ax.set_title('Численное решение модели Майнхардта',
                     color='#0066cc', fontsize=14, pad=20)

        self.canvas.draw()

    def show_stability_check(self, is_stable, limit):
        """Показывает результат проверки устойчивости и информационную рамку"""
        self.fig.clear()

        # Создаем два подграфика
        ax1 = self.fig.add_subplot(211)
        ax2 = self.fig.add_subplot(212)

        ax1.set_facecolor('#f0f8ff')
        ax2.set_facecolor('#f0f8ff')

        # Делаем окантовку графиков цветной
        for spine in ax1.spines.values():
            spine.set_edgecolor('#0066cc')
            spine.set_linewidth(1.5)

        for spine in ax2.spines.values():
            spine.set_edgecolor('#0066cc')
            spine.set_linewidth(1.5)

        if is_stable:
            status = "Условие устойчивости ВЫПОЛНЯЕТСЯ"
            color = '#0066cc'
        else:
            status = "Условие устойчивости НЕ ВЫПОЛНЯЕТСЯ!"
            color = '#ff6666'

        # Текст проверки устойчивости - БЕЗ внутренней рамки
        info_text = f"{status}\n\n" \
                    f"τ = {self.solver.tau:.6f}\n" \
                    f"h²/(2·max(D)) = {limit:.6f}\n\n" \
                    f"τ ≤ h²/(2·max(D_a, D_b, D_c))"

        ax1.text(0.5, 0.5, info_text, horizontalalignment='center',
                 verticalalignment='center', transform=ax1.transAxes,
                 fontsize=12, color=color)

        ax1.set_xticks([])
        ax1.set_yticks([])
        ax1.set_title('Проверка устойчивости схемы',
                      color='#0066cc', fontsize=12, pad=10)

        # Нижний блок - ТОЖЕ БЕЗ внутренней рамки
        ax2.text(0.5, 0.5, 'Для построения графиков решения\nвыберите команду из раздела "ВИЗУАЛИЗАЦИЯ"',
                 horizontalalignment='center', verticalalignment='center',
                 transform=ax2.transAxes, fontsize=11,
                 color='#0066cc')

        ax2.set_xticks([])
        ax2.set_yticks([])
        ax2.set_title('', color='#0066cc', fontsize=10)

        # Настройка размеров (нижний блок меньше)
        import matplotlib.gridspec as gridspec
        gs = gridspec.GridSpec(2, 1, height_ratios=[4, 1])
        ax1.set_position(gs[0].get_position(self.fig))
        ax2.set_position(gs[1].get_position(self.fig))

        self.fig.subplots_adjust(hspace=0.2)
        self.canvas.draw()

    def log_message(self, message):
        """Добавляет сообщение в информационное окно"""
        self.info_text.insert(tk.END, message + "\n")
        self.info_text.see(tk.END)

    def get_params(self):
        """Собирает все параметры из полей ввода"""
        params = {}
        for key, entry in self.param_entries.items():
            try:
                params[key] = float(entry.get())
            except ValueError:
                return None, f"Ошибка в параметре {key}"

        return params, None

    def get_grid_params(self):
        """Собирает параметры сетки"""
        grid_params = {}
        for key, entry in self.grid_entries.items():
            try:
                if key in ['n', 'm']:
                    grid_params[key] = int(entry.get())
                else:
                    grid_params[key] = float(entry.get())
            except ValueError:
                return None, f"Ошибка в параметре сетки {key}"

        return grid_params, None

    def get_init_params(self):
        """Собирает параметры начальных условий"""
        init_params = {}

        # Амплитуды
        for key, entry in self.init_entries.items():
            try:
                if key.startswith('b'):
                    init_params[key] = int(entry.get())
                else:
                    init_params[key] = float(entry.get())
            except ValueError:
                return None, f"Ошибка в параметре {key}"

        return init_params, None

    def start_calculation(self):
        """Запускает расчет в отдельном потоке"""
        # Получаем параметры
        params, error = self.get_params()
        if error:
            messagebox.showerror("Ошибка", error)
            return

        grid_params, error = self.get_grid_params()
        if error:
            messagebox.showerror("Ошибка", error)
            return

        init_params, error = self.get_init_params()
        if error:
            messagebox.showerror("Ошибка", error)
            return

        # Проверка положительности параметров
        if grid_params['n'] <= 0 or grid_params['m'] <= 0:
            messagebox.showerror("Ошибка", "n и m должны быть положительными числами")
            return

        if grid_params['T'] <= 0 or grid_params['L'] <= 0:
            messagebox.showerror("Ошибка", "T и L должны быть положительными числами")
            return

        # Блокируем кнопки
        self.calc_button.config(state=tk.DISABLED, bg='#cccccc')
        self.plot_2d_button.config(state=tk.DISABLED)
        self.plot_3d_button.config(state=tk.DISABLED)
        self.plot_surface_selected_button.config(state=tk.DISABLED)
        self.plot_surface_all_button.config(state=tk.DISABLED)
        self.export_button.config(state=tk.DISABLED)

        self.log_message("=" * 47)
        self.log_message("ПАРАМЕТРЫ МОДЕЛИ ОБНОВЛЕНЫ")
        self.log_message("=" * 47)

        # Создаем решатель
        self.solver = MainTaskNumericalSolver(params, init_params, grid_params)

        # Проверяем устойчивость
        is_stable, limit = self.solver.check_stability()
        self.show_stability_check(is_stable, limit)

        if not is_stable:
            self.log_message("=" * 47)
            self.log_message("!!! ОШИБКА: Условие устойчивости не выполняется !!!")
            self.log_message("=" * 47)
            self.log_message(f"τ = {self.solver.tau:.6f}")
            self.log_message(f"h²/(2·max(D)) = {limit:.6f}")
            self.log_message("")
            self.log_message("РЕКОМЕНДАЦИИ ДЛЯ УСТОЙЧИВОГО РАСЧЕТА:")
            self.log_message("  1. Увеличьте количество шагов по времени (m)")
            self.log_message("  2. Уменьшите конечное время (T)")
            self.log_message("  3. Уменьшите количество узлов по x (n)")
            self.log_message("")
            self.status_var.set("Ошибка: условие устойчивости нарушено")

            # Разблокируем кнопки
            self.calc_button.config(state=tk.NORMAL, bg='#4da6ff')

            # ВАЖНО: прерываем выполнение, НЕ запускаем расчёт
            return

        # Запускаем расчет в отдельном потоке (только если устойчиво)
        thread = threading.Thread(target=self.run_calculation)
        thread.daemon = True
        thread.start()

    def run_calculation(self):
        """Выполняет расчет в фоновом потоке"""
        success = self.solver.solve(self.log_message)

        # Обновляем интерфейс в главном потоке
        self.parent.after(0, self.calculation_finished, success)

    def calculation_finished(self, success):
        """Действия после завершения расчета"""
        if success:
            self.status_var.set("Расчет завершен")


            # ========== ЗНАЧЕНИЯ В КОНЕЧНЫЙ МОМЕНТ ==========
            a_final = self.solver.a[-1, :]
            b_final = self.solver.b[-1, :]
            c_final = self.solver.c[-1, :]

            self.log_message("\n1. ЗНАЧЕНИЯ В КОНЕЧНЫЙ МОМЕНТ ВРЕМЕНИ (t = %.2f):" % self.solver.T)
            self.log_message(f"   a = {a_final[0]:.10f}")
            self.log_message(f"   b = {b_final[0]:.10f}")
            self.log_message(f"   c = {c_final[0]:.10f}")

            # ========== ПРОВЕРКА НАЧАЛЬНЫХ УСЛОВИЙ ==========
            a_initial = self.solver.a[0, :]
            b_initial = self.solver.b[0, :]
            c_initial = self.solver.c[0, :]

            self.log_message("\n2. НАЧАЛЬНЫЕ УСЛОВИЯ (t = 0):")
            self.log_message(f"   a = {a_initial[0]:.10f}")
            self.log_message(f"   b = {b_initial[0]:.10f}")
            self.log_message(f"   c = {c_initial[0]:.10f}")


            # Активируем кнопки визуализации
            self.plot_2d_button.config(state=tk.NORMAL)
            self.plot_3d_button.config(state=tk.NORMAL)
            self.plot_surface_selected_button.config(state=tk.NORMAL)
            self.plot_surface_all_button.config(state=tk.NORMAL)
            self.export_button.config(state=tk.NORMAL)
            self.convergence_button.config(state=tk.NORMAL)
            self.save_to_excel(Dc_crit=76.797691171241346809)
        else:
            self.status_var.set("Ошибка расчета")
            self.log_message("ОШИБКА при расчете!")

        # Разблокируем кнопку расчета
        self.calc_button.config(state=tk.NORMAL, bg='#4da6ff')

    def parse_layers(self):
        """Парсит строку с номерами слоев"""
        try:
            layers_str = self.layers_var.get().strip()
            layers = [int(l.strip()) for l in layers_str.split(',')]

            if not self.solver:
                return None

            if len(layers) > 15:
                messagebox.showerror("Ошибка", "Можно выбрать не более 15 слоев")
                return None

            for layer in layers:
                if layer < 0 or layer > self.solver.m:
                    messagebox.showerror("Ошибка",
                                         f"Слой {layer} не существует.\n"
                                         f"Доступные слои: 0-{self.solver.m}")
                    return None

            return layers
        except ValueError:
            messagebox.showerror("Ошибка", "Некорректный формат слоев.\nИспользуйте: 0,1,2,3")
            return None

    def plot_2d(self):
        """Строит 2D график для выбранных слоев"""
        layers = self.parse_layers()
        if not layers:
            return

        substance = self.substance_var.get()

        self.fig.clear()
        ax = self.fig.add_subplot(111)
        ax.set_facecolor('#f0f8ff')
        ax.ticklabel_format(axis='y', useOffset=False, style='plain')

        colors = ['#0066cc', '#0099ff', '#00ccff', '#66d9ff', '#33adff']

        for idx, layer in enumerate(layers):
            x, u = self.solver.get_layer(substance, layer)
            if x is not None:
                color = colors[idx % len(colors)]
                ax.plot(x, u, 'o-', linewidth=2, markersize=4,
                        color=color, label=f'Слой {layer} (t={self.solver.t[layer]:.4f})')

        ax.set_xlabel('x', color='#0066cc', fontsize=12)
        ax.set_ylabel(f'{substance}(x,t)', color='#0066cc', fontsize=12)
        ax.set_title(f'Распределение {substance}(x,t) на отрезке [0, {self.solver.L}], слои {layers}',
                     color='#0066cc', fontsize=18)
        ax.grid(True, alpha=0.3, color='#66b3ff')
        ax.legend(facecolor='#f0f8ff', edgecolor='#0066cc')
        ax.tick_params(colors='#0066cc')

        self.canvas.draw()
        self.log_message(f"Построен 2D график для {substance}, слои {layers}")

    def plot_3d(self):
        """Строит 3D линейный график для выбранных слоев"""
        layers = self.parse_layers()
        if not layers:
            return

        substance = self.substance_var.get()

        self.fig.clear()
        ax = self.fig.add_subplot(111, projection='3d')

        colors = ['#0066cc', '#0099ff', '#00ccff', '#66d9ff', '#33adff']

        for idx, layer in enumerate(layers):
            x, u = self.solver.get_layer(substance, layer)
            if x is not None:
                color = colors[idx % len(colors)]
                t_layer = np.full_like(x, self.solver.t[layer])
                ax.plot(x, t_layer, u,
                        color=color,
                        linewidth=2.5,
                        marker='o',
                        markersize=4,
                        label=f'Слой {layer} (t={self.solver.t[layer]:.4f})')

        ax.set_xlabel('x', color='#0066cc', labelpad=15)
        ax.set_ylabel('t', color='#0066cc', labelpad=15)
        ax.set_zlabel(f'{substance}(x,t)', color='#0066cc', labelpad=15)
        ax.set_title(f'Распределение {substance}(x,t) на отрезке [0, {self.solver.L}], слои {layers}',
                     color='#0066cc', pad=20)
        ax.view_init(elev=35, azim=-140)

        legend = ax.legend(facecolor='#f0f8ff',
                           edgecolor='#0066cc',
                           loc='center left',
                           bbox_to_anchor=(-0.5, 0.8),
                           fontsize=10)

        ax.xaxis.pane.set_facecolor('#f0f8ff')
        ax.yaxis.pane.set_facecolor('#f0f8ff')
        ax.zaxis.pane.set_facecolor('#f0f8ff')
        ax.grid(True, alpha=0.3, color='#66b3ff')

        self.canvas.draw()
        self.log_message(f"Построен 3D график для {substance}, слои {layers}")

    def plot_surface(self):
        """Строит 3D поверхность для всех слоев с заданным шагом"""
        if not self.solver:
            return

        try:
            step = int(self.step_var.get())
            if step <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Ошибка", "Шаг должен быть положительным целым числом")
            return

        substance = self.substance_var.get()

        # Создаем список слоев с заданным шагом
        layers = list(range(0, self.solver.m + 1, step))


        if len(layers) < 2:
            messagebox.showerror("Ошибка", "Слишком мало слоев для построения поверхности")
            return

        self.fig.clear()
        ax = self.fig.add_subplot(111, projection='3d')

        # Подготавливаем данные для поверхности
        X, T = np.meshgrid(self.solver.x, self.solver.t[layers])

        if substance == 'a':
            U = self.solver.a[layers, :]
        elif substance == 'b':
            U = self.solver.b[layers, :]
        else:
            U = self.solver.c[layers, :]


        surf = ax.plot_surface(X, T, U, cmap='Blues', alpha=0.8)

        ax.set_xlabel('x', color='#0066cc', labelpad=15)
        ax.set_ylabel('t', color='#0066cc', labelpad=15)
        ax.set_zlabel(f'{substance}(x,t)', color='#0066cc', labelpad=15)
        ax.set_title(f'Поверхность {substance}(x,t), шаг={step}', color='#0066cc', pad=20)
        ax.view_init(elev=35, azim=-140)

        cbar = self.fig.colorbar(surf, ax=ax, shrink=0.6, aspect=20, pad=0.15)
        cbar.ax.tick_params(colors='#0066cc')
        cbar.set_label(f'{substance}(x,t)', color='#0066cc')

        ax.xaxis.pane.set_facecolor('#f0f8ff')
        ax.yaxis.pane.set_facecolor('#f0f8ff')
        ax.zaxis.pane.set_facecolor('#f0f8ff')

        self.canvas.draw()
        self.log_message(f"Построена поверхность для {substance}, шаг={step}, слоев={len(layers)}")

    def plot_surface_selected(self):
        """Строит 3D поверхность для выбранных пользователем слоёв"""
        if not self.solver:
            return

        # Получаем выбранные слои
        layers = self.parse_layers()
        if not layers:
            return

        if len(layers) < 2:
            messagebox.showerror("Ошибка", "Для построения поверхности нужно выбрать минимум 2 слоя")
            return

        substance = self.substance_var.get()

        # Сортируем слои по возрастанию
        layers.sort()

        self.log_message(f"Построение поверхности для слоев: {layers}")

        self.fig.clear()
        ax = self.fig.add_subplot(111, projection='3d')

        # Подготавливаем данные для поверхности
        X, T = np.meshgrid(self.solver.x, self.solver.t[layers])

        if substance == 'a':
            U = self.solver.a[layers, :]
        elif substance == 'b':
            U = self.solver.b[layers, :]
        else:
            U = self.solver.c[layers, :]

        # Рисуем поверхность
        surf = ax.plot_surface(X, T, U, cmap='Blues', alpha=0.8,
                               rstride=1, cstride=1,
                               edgecolor='white', linewidth=0.5)

        ax.set_xlabel('x', color='#0066cc', labelpad=15)
        ax.set_ylabel('t', color='#0066cc', labelpad=15)
        ax.set_zlabel(f'{substance}(x,t)', color='#0066cc', labelpad=15)
        ax.set_title(f'Поверхность {substance}(x,t) на отрезке [0, {self.solver.L}], слои {layers}',
                     color='#0066cc', pad=20)
        ax.view_init(elev=35, azim=-140)

        cbar = self.fig.colorbar(surf, ax=ax, shrink=0.6, aspect=20, pad=0.15)
        cbar.ax.tick_params(colors='#0066cc')
        cbar.set_label(f'{substance}(x,t)', color='#0066cc')

        ax.xaxis.pane.set_facecolor('#f0f8ff')
        ax.yaxis.pane.set_facecolor('#f0f8ff')
        ax.zaxis.pane.set_facecolor('#f0f8ff')

        self.canvas.draw()
        self.log_message(f"Построена поверхность для {substance}, слои {layers}")

    def export_data(self):
        """Экспортирует данные в файл"""
        if not self.solver:
            return

        try:
            step = int(self.step_var.get())
            if step <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Ошибка", "Шаг должен быть положительным целым числом")
            return

        # Создаем папку results если её нет
        if not os.path.exists("results"):
            os.makedirs("results")

        # Генерируем имя файла
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"results/Численное_решение_{timestamp}.txt"

        # Экспортируем
        success = self.solver.export_to_file(filename, step)

        if success:
            self.log_message(f"Данные экспортированы в {filename}")
            messagebox.showinfo("Успех", f"Данные экспортированы в файл:\n{filename}")
        else:
            messagebox.showerror("Ошибка", "Ошибка при экспорте данных")

    def check_convergence(self):
        """Проверка сеточной сходимости"""
        if not self.solver:
            messagebox.showwarning("Предупреждение", "Сначала выполните основной расчет!")
            return

        # Создаем диалоговое окно для выбора параметров
        settings_dialog = tk.Toplevel(self.parent)
        settings_dialog.title("Настройка проверки сходимости")
        settings_dialog.geometry("550x300")
        settings_dialog.transient(self.parent)
        settings_dialog.grab_set()

        main_frame = ttk.Frame(settings_dialog, style='Blue.TFrame')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        # Заголовок
        title_label = ttk.Label(main_frame, text="Настройка проверки сеточной сходимости",
                                style='Title.TLabel')
        title_label.pack(pady=(0, 15))

        # Поле ввода слоев
        layers_frame = ttk.Frame(main_frame, style='Blue.TFrame')
        layers_frame.pack(fill=tk.X, pady=5)

        ttk.Label(layers_frame, text="Слои для сравнения:", background='#cce5ff').pack(side=tk.LEFT, padx=(0, 5))
        layers_var = tk.StringVar(value="0,1,2,3,4,5,6,7,8,9,10")
        layers_entry = ttk.Entry(layers_frame, textvariable=layers_var, width=30)
        layers_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        layers_entry.focus()

        # Поле ввода количества расчетов
        iterations_frame = ttk.Frame(main_frame, style='Blue.TFrame')
        iterations_frame.pack(fill=tk.X, pady=5)

        ttk.Label(iterations_frame, text="Количество расчетов:", background='#cce5ff').pack(side=tk.LEFT, padx=(0, 5))
        iterations_var = tk.StringVar(value="5")
        iterations_entry = ttk.Entry(iterations_frame, textvariable=iterations_var, width=10)
        iterations_entry.pack(side=tk.LEFT)

        # Подсказки
        hints_frame = ttk.Frame(main_frame, style='Blue.TFrame')
        hints_frame.pack(fill=tk.X, pady=(5, 10))

        hints_text = "• Слои по времени: введите слои для сравнения (через запятую)\n" \
                     "• Количество расчетов: от 2 до 5 (каждый следующий расчет с удвоенной сеткой)"

        hint_label = ttk.Label(hints_frame, text=hints_text, background='#cce5ff',
                               font=('Arial', 8), foreground='#666666', justify=tk.LEFT)
        hint_label.pack(anchor=tk.W)

        # Информация о доступных слоях
        max_layer = self.solver.m
        info_label = ttk.Label(main_frame, text=f"Доступные слои: от 0 до {max_layer}",
                               background='#cce5ff', font=('Arial', 9), foreground='#0066cc')
        info_label.pack(pady=(0, 10))

        def start_check():
            try:
                # Парсим слои
                layers_str = layers_var.get().strip()
                layers_list = [int(l.strip()) for l in layers_str.split(',')]

                if len(layers_list) > 15:
                    messagebox.showerror("Ошибка", "Можно выбрать не более 15 слоев")
                    return

                for layer in layers_list:
                    if layer < 0 or layer > max_layer:
                        messagebox.showerror("Ошибка", f"Слой {layer} не существует.\nДоступные слои: 0-{max_layer}")
                        return

                # Парсим количество расчетов
                max_iterations = int(iterations_var.get())
                if max_iterations < 2 or max_iterations > 5:
                    raise ValueError

                settings_dialog.destroy()

                # Получаем базовые параметры
                params, _ = self.get_params()
                init_params, _ = self.get_init_params()
                grid_params, _ = self.get_grid_params()

                T = grid_params['T']
                L = grid_params['L']
                base_n = self.solver.n
                base_m = self.solver.m

                self.log_message("=" * 50)
                self.log_message("ПРОВЕРКА СЕТОЧНОЙ СХОДИМОСТИ")
                self.log_message("=" * 50)
                self.log_message(f"Слои для сравнения: {layers_list}")
                self.log_message(f"Количество расчетов: {max_iterations}")

                # Запускаем в отдельном потоке
                thread = threading.Thread(target=self.run_convergence_check,
                                          args=(params, init_params, T, L, base_n, base_m, layers_list, max_iterations))
                thread.daemon = True
                thread.start()

            except ValueError:
                messagebox.showerror("Ошибка", "Некорректный формат слоев или количества расчетов")

        # Кнопка запуска
        start_button = tk.Button(main_frame, text="ЗАПУСТИТЬ ПРОВЕРКУ",
                                 command=start_check, bg='#4da6ff', fg='white',
                                 font=('Arial', 10, 'bold'), width=25)
        start_button.pack(pady=8, fill=tk.X, padx=30)

        layers_entry.bind('<Return>', lambda e: start_check())
        iterations_entry.bind('<Return>', lambda e: start_check())

    def run_convergence_check(self, params, init_params, T, L, base_n, base_m, layers, max_iterations):
        """Запускает проверку сходимости в фоновом потоке"""
        try:
            from convergence_checker import ConvergenceChecker
            from main_model import MainTaskNumericalSolver

            checker = ConvergenceChecker(self, MainTaskNumericalSolver)

            # Запускаем проверку
            results, iteration_details = checker.check_convergence(
                base_n=base_n,
                base_m=base_m,
                T=T,
                L=L,  # L передается здесь
                params=params,
                init_params=init_params,
                layers=layers,
                max_iterations=max_iterations
            )

            if results:
                # Передаем T и L в функцию отображения таблицы
                self.parent.after(0, lambda: self.plot_convergence_table(results, iteration_details, T, L))
                self.log_message(f"Выполнено расчетов: {len(results)}")

        except Exception as e:
            self.log_message(f"Ошибка при проверке сходимости: {str(e)}")
            import traceback
            self.log_message(traceback.format_exc())

    def plot_convergence_table(self, results, iteration_details, T, L):
        """Показывает три таблицы результатов проверки сходимости (для a, b, c)"""
        self.fig.clear()

        # Создаем три подграфика друг над другом
        ax1 = self.fig.add_subplot(311)
        ax2 = self.fig.add_subplot(312)
        ax3 = self.fig.add_subplot(313)

        for ax in [ax1, ax2, ax3]:
            ax.set_facecolor('white')
            ax.axis('off')

        # Создаем информацию о слоях для отображения
        layers_info = []
        if iteration_details and len(iteration_details) > 0:
            first_detail = iteration_details[0]
            actual_layers = [lc['fine_layer'] for lc in first_detail['layer_comparisons']]
            layers_info = actual_layers

        # Заголовки таблицы (без h и τ)
        headers = ['№ расчета', 'n', 'm', 'Макс. погрешность', 'Порядок сходимости']

        # Функция для сбора погрешностей по конкретному веществу
        def get_max_diffs_for_substance(iteration_details, substance_idx):
            """substance_idx: 0 - a, 1 - b, 2 - c"""
            max_diffs = []
            for i in range(len(iteration_details) - 1):
                # Берем данные сравнения между i и i+1 сетками
                # В iteration_details[i+1] хранятся сравнения с предыдущей сеткой
                detail = iteration_details[i + 1]
                max_diff = 0.0
                for layer_comp in detail['layer_comparisons']:
                    if substance_idx == 0:  # a
                        diff = layer_comp['max_diff_a']
                    elif substance_idx == 1:  # b
                        diff = layer_comp['max_diff_b']
                    else:  # c
                        diff = layer_comp['max_diff_c']
                    if diff > max_diff:
                        max_diff = diff
                max_diffs.append(max_diff)
            return max_diffs

        # Функция для построения одной таблицы
        def build_table_data(iteration_details, max_diffs, substance_name):
            table_data = []
            for i, detail in enumerate(iteration_details, 1):
                if i == 1:
                    max_diff_str = "—"
                elif i - 2 < len(max_diffs):
                    max_diff_str = f"{max_diffs[i - 2]:.2e}"
                else:
                    max_diff_str = "—"

                if i == 1:
                    order_str = "—"
                else:
                    # Используем индивидуальный порядок для вещества
                    if substance_name == 'a':
                        rate = detail.get('convergence_rate_a')
                    elif substance_name == 'b':
                        rate = detail.get('convergence_rate_b')
                    else:  # 'c'
                        rate = detail.get('convergence_rate_c')

                    if rate is not None and abs(rate) < 0.01:
                        order_str = "—"
                    elif rate is not None:
                        order_str = f"{rate:.3f}"
                    else:
                        order_str = "—"

                table_data.append([i, detail['n'], detail['m'], max_diff_str, order_str])
            return table_data

        # Получаем погрешности для каждого вещества
        max_diffs_a = get_max_diffs_for_substance(iteration_details, 0)
        max_diffs_b = get_max_diffs_for_substance(iteration_details, 1)
        max_diffs_c = get_max_diffs_for_substance(iteration_details, 2)

        # Строим таблицы
        table_data_a = build_table_data(iteration_details, max_diffs_a, 'a')
        table_data_b = build_table_data(iteration_details, max_diffs_b, 'b')
        table_data_c = build_table_data(iteration_details, max_diffs_c, 'c')

        # Создаем таблицы на соответствующих осях
        table_a = ax1.table(cellText=table_data_a, colLabels=headers,
                            cellLoc='center', loc='center',
                            colColours=['#e6f2ff'] * len(headers))
        table_b = ax2.table(cellText=table_data_b, colLabels=headers,
                            cellLoc='center', loc='center',
                            colColours=['#e6f2ff'] * len(headers))
        table_c = ax3.table(cellText=table_data_c, colLabels=headers,
                            cellLoc='center', loc='center',
                            colColours=['#e6f2ff'] * len(headers))

        # Настройка внешнего вида таблиц
        for table in [table_a, table_b, table_c]:
            table.auto_set_font_size(False)
            table.set_fontsize(10)
            table.scale(1.2, 1.6)

            # Стилизация заголовков
            for j in range(len(headers)):
                table[(0, j)].set_facecolor('#e6f2ff')
                table[(0, j)].set_text_props(weight='bold', color='#0066cc', fontsize=10)

        # Заголовки для каждой таблицы (названия веществ)
        if layers_info:
            title_text = f'Проверка сеточной сходимости\n(сравнение слоев: {layers_info})'
        else:
            title_text = 'Проверка сеточной сходимости'

        self.fig.suptitle(title_text, color='#0066cc', fontsize=14, y=0.98)

        ax1.set_title('\n\n Вещество a(x,t)', color='#0066cc', fontsize=12, loc='center')
        ax2.set_title('\n\n Вещество b(x,t)', color='#0066cc', fontsize=12, loc='center')
        ax3.set_title('\n\n Вещество c(x,t)', color='#0066cc', fontsize=12, loc='center')

        # Удален блок с пояснениями (explanations)

        # Настройка отступов между подграфиками
        self.fig.subplots_adjust(hspace=0.35, top=0.86, bottom=0.12)

        self.canvas.draw()
        self.log_message(f"Построены три таблицы сходимости для a, b, c (слои: {layers_info})")

    def save_to_excel(self, Dc_crit=76.8):
        """Автоматически сохраняет результаты в Excel-файл"""
        if not self.solver or not self.solver.calculation_done:
            return

        try:
            import openpyxl
        except ImportError:
            self.log_message("!!! Установите openpyxl: pip install openpyxl")
            return

        filename = "results/таблица_результатов.xlsx"
        os.makedirs("results", exist_ok=True)

        D_c = self.solver.params['Dc']
        layer_idx = self.solver.m

        a_min = float(np.min(self.solver.a[-1, :]))
        a_max = float(np.max(self.solver.a[-1, :]))
        b_min = float(np.min(self.solver.b[-1, :]))
        b_max = float(np.max(self.solver.b[-1, :]))
        c_min = float(np.min(self.solver.c[-1, :]))
        c_max = float(np.max(self.solver.c[-1, :]))

        razmah_a = a_max - a_min
        razmah_b = b_max - b_min
        razmah_c = c_max - c_min
        # Оценка времени формирования структуры (ищем с начала)
        # Время формирования: идём с конца, сравниваем с последним слоем
        form_time = self.solver.T
        tol = 2e-4
        last_a = self.solver.a[-1, :]
        last_b = self.solver.b[-1, :]
        last_c = self.solver.c[-1, :]

        for j in range(self.solver.m - 1, -1, -1):
            diff_a = np.max(np.abs(self.solver.a[j, :] - last_a))
            diff_b = np.max(np.abs(self.solver.b[j, :] - last_b))
            diff_c = np.max(np.abs(self.solver.c[j, :] - last_c))
            max_diff = max(diff_a, diff_b, diff_c)

            if max_diff > tol:
                form_time = self.solver.t[min(j + 1, self.solver.m)]
                break


        if os.path.exists(filename):
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            next_row = ws.max_row + 1
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = ['D_c', 'D_c - Dc_крит', 'sign(D_c-Dc_крит)',
                       'a(min)', 'a(max)', 'b(min)', 'b(max)', 'c(min)', 'c(max)',
                       'm', 'τ', 'слой', 'время форм', 'размах a', 'размах b', 'размах c']
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
            next_row = 2

        ws.cell(row=next_row, column=1, value=D_c)
        ws.cell(row=next_row, column=2, value=D_c - Dc_crit)
        diff = D_c - Dc_crit
        ws.cell(row=next_row, column=3, value=1 if diff > 0 else (-1 if diff < 0 else 0))
        ws.cell(row=next_row, column=4, value=a_min)
        ws.cell(row=next_row, column=5, value=a_max)
        ws.cell(row=next_row, column=6, value=b_min)
        ws.cell(row=next_row, column=7, value=b_max)
        ws.cell(row=next_row, column=8, value=c_min)
        ws.cell(row=next_row, column=9, value=c_max)
        ws.cell(row=next_row, column=10, value=layer_idx)
        ws.cell(row=next_row, column=11, value=form_time)
        ws.cell(row=next_row, column=12, value=razmah_a)
        ws.cell(row=next_row, column=13, value=razmah_b)
        ws.cell(row=next_row, column=14, value=razmah_c)
        ws.cell(row=next_row, column=15, value=self.solver.m)
        ws.cell(row=next_row, column=16, value=self.solver.tau)

        for col in range(1, 17):
            ws.cell(row=next_row, column=col).number_format = '0.000000'

        wb.save(filename)
        self.log_message(f"✓ Сохранено в Excel: строка {next_row}, Dc={D_c}")

    def find_intersections_all_time(self):
        """Находит точки пересечения a=b, b=c, a=c для всех временных слоев"""
        if not self.solver or not self.solver.calculation_done:
            self.log_message("Расчет еще не выполнен или не завершен")
            return

        self.log_message("=" * 70)
        self.log_message("ПОИСК ТОЧЕК ПЕРЕСЕЧЕНИЯ")
        self.log_message("=" * 70)

        x = self.solver.x
        intersections = {'ab': [], 'bc': [], 'ac': []}

        total_layers = self.solver.m + 1
        self.log_message(f"Всего временных слоев: {total_layers}")

        for j in range(total_layers):
            t_val = self.solver.t[j]
            a_layer = self.solver.a[j, :]
            b_layer = self.solver.b[j, :]
            c_layer = self.solver.c[j, :]

            diff_ab = a_layer - b_layer
            diff_bc = b_layer - c_layer
            diff_ac = a_layer - c_layer

            # Поиск a=b
            for i in range(len(x) - 1):
                if diff_ab[i] * diff_ab[i + 1] <= 0:
                    x1, x2 = x[i], x[i + 1]
                    y1, y2 = diff_ab[i], diff_ab[i + 1]
                    if abs(y2 - y1) > 1e-12:
                        t_interp = -y1 / (y2 - y1)
                        x_inter = x1 + t_interp * (x2 - x1)
                        y_inter = a_layer[i] + t_interp * (a_layer[i + 1] - a_layer[i])
                        intersections['ab'].append((t_val, x_inter, y_inter))

            # Поиск b=c
            for i in range(len(x) - 1):
                if diff_bc[i] * diff_bc[i + 1] <= 0:
                    x1, x2 = x[i], x[i + 1]
                    y1, y2 = diff_bc[i], diff_bc[i + 1]
                    if abs(y2 - y1) > 1e-12:
                        t_interp = -y1 / (y2 - y1)
                        x_inter = x1 + t_interp * (x2 - x1)
                        y_inter = b_layer[i] + t_interp * (b_layer[i + 1] - b_layer[i])
                        intersections['bc'].append((t_val, x_inter, y_inter))

            # Поиск a=c
            for i in range(len(x) - 1):
                if diff_ac[i] * diff_ac[i + 1] <= 0:
                    x1, x2 = x[i], x[i + 1]
                    y1, y2 = diff_ac[i], diff_ac[i + 1]
                    if abs(y2 - y1) > 1e-12:
                        t_interp = -y1 / (y2 - y1)
                        x_inter = x1 + t_interp * (x2 - x1)
                        y_inter = a_layer[i] + t_interp * (a_layer[i + 1] - a_layer[i])
                        intersections['ac'].append((t_val, x_inter, y_inter))

        # Выводим результаты
        self.log_message(f"\nНайдено пересечений a=b: {len(intersections['ab'])}")
        self.log_message(f"Найдено пересечений b=c: {len(intersections['bc'])}")
        self.log_message(f"Найдено пересечений a=c: {len(intersections['ac'])}")

        # Для вашего случая a=b будет много, b=c и a=c - ноль
        self.log_message("=" * 70)


# ====================== ЗАПУСК ПРИЛОЖЕНИЯ ======================

def main():
    root = tk.Tk()
    app = ActivatorInhibitorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()